#!/usr/bin/env python3
"""
Deterministic generator for Excel parser fixtures.

Regenerates the W10 (merged cells), W11 (image-only), and W40 (additional
messy workbooks) fixtures under fixtures/parsers/. Existing baseline fixtures
are NOT touched by this script.

Usage:
    python scripts/gen-parser-fixtures.py

The produced workbooks contain no timestamps or random values so the parsed
output stays deterministic across runs.
"""

import io
import struct
import zlib
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures" / "parsers"


def write_rows(ws, rows):
    for row in rows:
        ws.append(list(row))


def make_png_bytes(width: int = 24, height: int = 24) -> bytes:
    """Build a tiny deterministic solid-color PNG without external assets."""
    def chunk(tag: bytes, data: bytes) -> bytes:
        return (
            struct.pack(">I", len(data))
            + tag
            + data
            + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF)
        )

    signature = b"\x89PNG\r\n\x1a\n"
    ihdr = struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)
    # One filter byte (0) per scanline, then RGB pixels (steel blue).
    raw = b"".join(b"\x00" + b"\x46\x82\xb4" * width for _ in range(height))
    idat = zlib.compress(raw, 9)
    return signature + chunk(b"IHDR", ihdr) + chunk(b"IDAT", idat) + chunk(b"IEND", b"")


# ---------------------------------------------------------------------------
# W10 - Merged-cell workbooks
# ---------------------------------------------------------------------------

def gen_merged_header_rent_roll():
    """Rent roll with a merged title banner AND a single header label merged
    horizontally across the two rent columns.

    The "Market Rent" header cell is merged across D2:E2, so without unmerge +
    forward-fill the Current Rent column header reads as blank/NaN (and the
    current-rent mapping is silently lost). After forward-fill the merged value
    propagates into both rent columns, producing TWO identically named "Market
    Rent" columns. That is genuinely ambiguous, so the parser must surface a
    candidate-review WARNING rather than silently mis-map current vs market rent.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged RR"

    # Title banner merged across all 6 columns (row 1) - benign, skipped by
    # header detection because it is a single populated cell.
    ws["A1"] = "Hillcrest Apartments - Rent Roll"
    ws.merge_cells("A1:F1")

    # Header row (row 2). Merge "Market Rent" across the two rent columns.
    ws["A2"] = "Unit"
    ws["B2"] = "Unit Type"
    ws["C2"] = "SqFt"
    ws["D2"] = "Market Rent"
    ws["F2"] = "Status"
    ws.merge_cells("D2:E2")

    rows = [
        ("101", "1x1", 720, 1650, 1575, "Occupied"),
        ("102", "1x1", 720, 1650, 0, "Vacant"),
        ("201", "2x2", 1050, 2250, 2175, "Occupied"),
        ("202", "2x2", 1050, 2250, 2200, "Occupied"),
    ]
    for row in rows:
        ws.append(list(row))

    wb.save(FIXTURES / "rent-roll-merged-headers.xlsx")


def gen_merged_label_column_rent_roll():
    """Rent roll where a merged section banner spans the full table width and
    the unit-type column uses vertically merged cells (a single value covering
    several stacked rows). Forward-fill must propagate the merged label down."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged Label"

    ws["A1"] = "Building A - Units"
    ws.merge_cells("A1:F1")

    ws["A2"] = "Unit"
    ws["B2"] = "Floor Plan"
    ws["C2"] = "SqFt"
    ws["D2"] = "Market Rent"
    ws["E2"] = "Current Rent"
    ws["F2"] = "Status"

    rows = [
        ("101", "1x1", 700, 1600, 1550, "Occupied"),
        ("102", None, 700, 1600, 1525, "Occupied"),   # floor plan merged from above
        ("201", "2x2", 1050, 2250, 2200, "Occupied"),
        ("202", None, 1050, 2250, 0, "Vacant"),       # floor plan merged from above
    ]
    for row in rows:
        ws.append(list(row))

    # Vertically merge the floor-plan column for the 1x1 group (B3:B4) and the
    # 2x2 group (B5:B6). openpyxl keeps only the top-left value populated.
    ws.merge_cells("B3:B4")
    ws.merge_cells("B5:B6")

    wb.save(FIXTURES / "rent-roll-merged-label-column.xlsx")


# ---------------------------------------------------------------------------
# W11 - Image-only workbook
# ---------------------------------------------------------------------------

def gen_image_only_workbook():
    """Workbook whose target sheet contains an embedded image and essentially
    no extractable tabular text (a scanned rent roll exported as a picture)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scanned RR"

    png_path = FIXTURES / "_scanned-rent-roll.png"
    png_path.write_bytes(make_png_bytes())
    img = XLImage(str(png_path))
    img.anchor = "A1"
    ws.add_image(img)

    wb.save(FIXTURES / "rent-roll-image-only.xlsx")
    # The PNG is only needed at workbook-build time; openpyxl has embedded a
    # copy inside the xlsx. Remove the loose helper so the fixture dir stays
    # to .xlsx files only.
    png_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# W40 - Additional messy fixtures
# ---------------------------------------------------------------------------

def gen_trailing_notes_rent_roll():
    """Real table followed by free-text disclaimer/notes rows after the data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trailing Notes"
    write_rows(ws, [
        ("Unit", "Unit Type", "SqFt", "Market Rent", "Current Rent", "Status"),
        ("101", "1x1", 720, 1650, 1600, "Occupied"),
        ("102", "1x1", 720, 1650, 1575, "Occupied"),
        ("201", "2x2", 1050, 2250, 0, "Vacant"),
        ("202", "2x2", 1050, 2250, 2225, "Occupied"),
        (None, None, None, None, None, None),
        ("Notes:", None, None, None, None, None),
        ("Rents reflect April 2026 billing.", None, None, None, None, None),
        ("Prepared by Property Management Co.", None, None, None, None, None),
    ])
    wb.save(FIXTURES / "rent-roll-trailing-notes.xlsx")


def gen_currency_symbols_rent_roll():
    """Rents stored as text with currency symbols and thousands separators."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Currency"
    write_rows(ws, [
        ("Unit", "Unit Type", "SqFt", "Market Rent", "Current Rent", "Status"),
        ("101", "1x1", 720, "$1,650", "$1,600", "Occupied"),
        ("102", "1x1", 720, "$1,650", "$0", "Vacant"),
        ("201", "2x2", "1,050", "$2,250", "$2,200", "Occupied"),
        ("202", "2x2", "1,050", "$2,250", "$2,175", "Occupied"),
    ])
    wb.save(FIXTURES / "rent-roll-currency-symbols.xlsx")


def gen_subtotal_rows_rent_roll():
    """Multiple per-building Subtotal rows plus a Grand Total that must all be
    excluded from the unit count and aggregates."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subtotals"
    write_rows(ws, [
        ("Unit", "Unit Type", "SqFt", "Market Rent", "Current Rent", "Status"),
        ("101", "1x1", 700, 1600, 1550, "Occupied"),
        ("102", "1x1", 700, 1600, 1525, "Occupied"),
        ("Subtotal", None, 1400, 3200, 3075, None),
        ("201", "2x2", 1050, 2250, 2200, "Occupied"),
        ("202", "2x2", 1050, 2250, 0, "Vacant"),
        ("Subtotal", None, 2100, 4500, 2200, None),
        ("Grand Total", None, 3500, 7700, 5275, None),
    ])
    wb.save(FIXTURES / "rent-roll-subtotal-rows.xlsx")


def gen_synonym_headers_t12():
    """T12 using header/label synonyms not previously covered: a 'Description'
    line-item column, an 'EGI' revenue label, and 'OpEx Total'/'NOI' rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T12 Synonyms"
    write_rows(ws, [
        ("Description", "Q1", "Q2", "Q3", "Q4", "Trailing 12"),
        ("Effective Gross Income", 50000, 51000, 52000, 53000, 624000),
        ("Total OpEx", 20000, 21000, 20500, 21500, 252000),
        ("Net Operating Income", 30000, 30000, 31500, 31500, 372000),
    ])
    wb.save(FIXTURES / "t12-synonym-headers.xlsx")


def gen_currency_symbols_t12():
    """T12 with currency-formatted text totals and a leading title banner."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T12 Currency"
    write_rows(ws, [
        ("Riverside Gardens T12", None, None),
        ("Account", "Monthly Avg", "Annual Total"),
        ("Effective Gross Income", "$104,000", "$1,248,000"),
        ("Total Operating Expenses", "$42,500", "$510,000"),
        ("Net Operating Income", "$61,500", "$738,000"),
    ])
    wb.save(FIXTURES / "t12-currency-symbols.xlsx")


def main():
    FIXTURES.mkdir(parents=True, exist_ok=True)
    # W10
    gen_merged_header_rent_roll()
    gen_merged_label_column_rent_roll()
    # W11
    gen_image_only_workbook()
    # W40
    gen_trailing_notes_rent_roll()
    gen_currency_symbols_rent_roll()
    gen_subtotal_rows_rent_roll()
    gen_synonym_headers_t12()
    gen_currency_symbols_t12()
    print("Generated parser fixtures in", FIXTURES)


if __name__ == "__main__":
    main()
