#!/usr/bin/env python3
"""
Excel Parser for CRE Documents
Extracts structured data from rent rolls and T12 financials.

Usage:
    python parse_excel.py <file_path> [--type rent_roll|t12|auto]

Output:
    JSON to stdout
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List

try:
    import pandas as pd
    import openpyxl
except ImportError:
    print(json.dumps({
        "error": "Missing dependencies. Install with: pip install pandas openpyxl",
        "success": False
    }))
    sys.exit(1)


def normalize_text(value: Any) -> str:
    """Normalize labels while keeping word boundaries for fuzzy matching."""
    if pd.isna(value):
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def compact_text(value: Any) -> str:
    """Normalize labels for exact-ish comparisons that ignore punctuation."""
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value))


def read_sheet_metadata(file_path: str) -> Dict[str, Dict[str, Any]]:
    """Collect per-sheet merged-cell ranges and image counts via openpyxl.

    Returned shape:
        {sheet_name: {"merged": [(min_row, min_col, max_row, max_col), ...],
                       "image_count": int,
                       "text_cell_count": int}}
    Rows/cols are 1-based Excel coordinates (matching openpyxl).
    """
    metadata: Dict[str, Dict[str, Any]] = {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception:
        return metadata

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged = []
        for merged_range in list(ws.merged_cells.ranges):
            merged.append((merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col))

        # Count cells that carry extractable text (ignore pure whitespace).
        text_cell_count = 0
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                if isinstance(value, str):
                    if value.strip():
                        text_cell_count += 1
                else:
                    text_cell_count += 1

        image_count = 0
        try:
            image_count = len(getattr(ws, "_images", []) or [])
        except Exception:
            image_count = 0

        metadata[sheet_name] = {
            "merged": merged,
            "image_count": image_count,
            "text_cell_count": text_cell_count,
        }

    try:
        wb.close()
    except Exception:
        pass
    return metadata


def fill_merged_cells(raw_df: pd.DataFrame, merged_ranges: List[tuple]) -> pd.DataFrame:
    """Forward-fill the top-left value of each merged range across the range.

    openpyxl/pandas only populate the top-left cell of a merged region; the
    remaining cells are blank/NaN, which silently degrades header detection and
    column/label mapping. We propagate the anchor value across every cell in the
    merged span so a merged header row maps correctly and merged label columns
    inherit their value.
    """
    if not merged_ranges:
        return raw_df

    # Use object dtype so propagating a string anchor into an otherwise-numeric
    # column does not trigger pandas dtype-incompatibility warnings/errors.
    # Downstream parsing normalizes/cleans values regardless of dtype.
    filled = raw_df.astype(object).copy()
    n_rows = filled.shape[0]
    n_cols = filled.shape[1]
    for (min_row, min_col, max_row, max_col) in merged_ranges:
        # Convert 1-based Excel coordinates to 0-based DataFrame indices.
        top = min_row - 1
        left = min_col - 1
        if top < 0 or left < 0 or top >= n_rows or left >= n_cols:
            continue
        anchor = filled.iat[top, left]
        if pd.isna(anchor):
            continue

        is_horizontal = max_col > min_col
        is_vertical = max_row > min_row

        # A purely horizontal single-row merge that has NO other data in its row
        # is a full-width title/section banner, not a column header group.
        # Forward-filling it would fabricate a fake multi-column header row and
        # corrupt header detection, so skip it (the banner stays a single cell,
        # which header detection already ignores).
        if is_horizontal and not is_vertical:
            row_values = [
                filled.iat[top, c]
                for c in range(n_cols)
                if not (left <= c < max_col)
            ]
            has_other_data = any(
                not pd.isna(value) and str(value).strip() for value in row_values
            )
            if not has_other_data:
                continue

        for r in range(top, min(max_row, n_rows)):
            for c in range(left, min(max_col, n_cols)):
                if r == top and c == left:
                    continue
                if pd.isna(filled.iat[r, c]):
                    filled.iat[r, c] = anchor
    return filled


def detect_document_type(df: pd.DataFrame, filename: str) -> str:
    """Detect if document is rent roll or T12 based on content."""
    filename_lower = filename.lower()

    # Check filename first
    if any(kw in filename_lower for kw in ['rent', 'roll', 'unit', 'roster']):
        return 'rent_roll'
    if any(kw in filename_lower for kw in ['t12', 'trailing', 'income', 'operating', 'financial']):
        return 't12'

    # Check column headers
    columns_str = ' '.join(normalize_text(c) for c in df.columns)

    rent_roll_indicators = ['unit', 'apt', 'tenant', 'lease', 'sqft', 'sf']
    t12_indicators = ['revenue', 'expense', 'noi', 'income', 'vacancy']

    rent_score = sum(1 for ind in rent_roll_indicators if ind in columns_str)
    t12_score = sum(1 for ind in t12_indicators if ind in columns_str)

    if rent_score > t12_score:
        return 'rent_roll'
    elif t12_score > rent_score:
        return 't12'

    return 'unknown'


def find_header_row(df: pd.DataFrame, max_rows: int = 15) -> int:
    """Find the row containing column headers."""
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i]
        non_null = row.dropna()
        if len(non_null) >= 3:
            row_str = ' '.join(normalize_text(v) for v in non_null)
            # Look for common header indicators
            if any(kw in row_str for kw in ['unit', 'rent', 'type', 'status', 'line item', 'account', 'annual', 't12']):
                return i
    return 0


def prepare_table(raw_df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Convert a raw worksheet into a parsed table while preserving Excel row/column provenance."""
    header_row = find_header_row(raw_df)
    raw_headers = [
        str(value).strip() if not pd.isna(value) else None
        for value in raw_df.iloc[header_row]
    ]

    # Detect duplicate non-empty header labels. This happens when a single
    # merged header cell has been forward-filled across multiple columns, which
    # makes the resulting column mapping genuinely ambiguous (e.g. two columns
    # both labelled "Market Rent"). We disambiguate the column names so each is
    # addressable, but record the ambiguity so parsers can surface a warning
    # instead of silently mis-mapping.
    seen: Dict[str, int] = {}
    ambiguous_headers: List[str] = []
    headers: List[str] = []
    for index, value in enumerate(raw_headers):
        if value is None or value == "":
            headers.append(f"Column {index + 1}")
            continue
        normalized = normalize_text(value)
        if normalized and normalized in seen:
            seen[normalized] += 1
            if value not in ambiguous_headers:
                ambiguous_headers.append(value)
            headers.append(f"{value} ({seen[normalized]})")
        else:
            seen[normalized] = 0
            headers.append(value)

    data = raw_df.iloc[header_row + 1:].copy()
    data.columns = headers
    data["__excel_row_number"] = data.index + 1
    data = data.reset_index(drop=True)
    data.attrs["sheet_name"] = sheet_name
    data.attrs["header_row_number"] = header_row + 1
    data.attrs["column_positions"] = {str(header): index + 1 for index, header in enumerate(headers)}
    data.attrs["ambiguous_headers"] = ambiguous_headers
    return data


def sheet_score(raw_df: pd.DataFrame, doc_type: str) -> int:
    text = " ".join(
        normalize_text(value)
        for value in raw_df.head(20).fillna("").to_numpy().flatten().tolist()
        if normalize_text(value)
    )
    if doc_type == "rent_roll":
        terms = ["unit", "apt", "floor plan", "bed bath", "rent", "sqft", "occupied", "tenant"]
    elif doc_type == "t12":
        terms = ["line item", "account", "revenue", "income", "expense", "noi", "annual", "t12", "trailing"]
    else:
        terms = ["unit", "rent", "revenue", "income", "expense", "noi"]
    return sum(1 for term in terms if term in text)


def choose_sheet(workbook: Dict[str, pd.DataFrame], doc_type: str) -> tuple[str, pd.DataFrame]:
    """Pick the worksheet most likely to contain the requested document type."""
    scored = sorted(
        ((sheet_score(raw_df, doc_type), sheet_name, raw_df) for sheet_name, raw_df in workbook.items()),
        key=lambda item: item[0],
        reverse=True,
    )
    return scored[0][1], scored[0][2]


def label_matches(label: Any, keyword: str) -> bool:
    normalized_label = normalize_text(label)
    normalized_keyword = normalize_text(keyword)
    if not normalized_label or not normalized_keyword:
        return False
    if normalized_label == normalized_keyword:
        return True
    if compact_text(label) == compact_text(keyword):
        return True
    keyword_tokens = normalized_keyword.split()
    return all(token in normalized_label.split() for token in keyword_tokens)


def find_column(columns: List[Any], keywords: List[str], used: Optional[set] = None) -> Optional[Any]:
    used = used or set()
    available = [column for column in columns if column not in used and not str(column).startswith("__")]
    for keyword in keywords:
        for column in available:
            if normalize_text(column) == normalize_text(keyword) or compact_text(column) == compact_text(keyword):
                return column
    for keyword in keywords:
        for column in available:
            if label_matches(column, keyword):
                return column
    return None


def excel_column_letter(df: pd.DataFrame, column: Any) -> Optional[str]:
    positions = df.attrs.get("column_positions", {})
    position = positions.get(str(column))
    if not position:
        return None
    return openpyxl.utils.get_column_letter(position)


def excel_location(df: pd.DataFrame, row_number: Optional[int], column: Optional[Any], description: str) -> Dict[str, Any]:
    location = {
        "sheet": df.attrs.get("sheet_name", "Source Data"),
        "description": description,
    }
    if row_number:
        location["row"] = int(row_number)
    column_letter = excel_column_letter(df, column) if column is not None else None
    if column_letter:
        location["column"] = column_letter
    return location


def empty_row(row: pd.Series) -> bool:
    values = [value for key, value in row.items() if key != "__excel_row_number"]
    return all(pd.isna(value) or str(value).strip() == "" for value in values)


def total_like(value: Any) -> bool:
    label = normalize_text(value)
    return bool(re.search(r"\b(grand total|total|subtotal|summary|average|avg)\b", label))


def clean_currency(value: Any) -> Optional[float]:
    """Convert currency string to float."""
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


def clean_percentage(value: Any) -> Optional[float]:
    """Convert percentage string to decimal."""
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value) if value <= 1 else float(value) / 100
    s = str(value).replace('%', '').strip()
    try:
        v = float(s)
        return v / 100 if v > 1 else v
    except ValueError:
        return None


def normalize_unit_type(value: Any) -> str:
    """Normalize unit type to standard format."""
    if pd.isna(value):
        return "Unknown"

    s = str(value).upper().strip()

    # Handle common patterns
    patterns = [
        (r'(\d)[xX](\d)', r'\1BR/\2BA'),  # 2x2 -> 2BR/2BA
        (r'(\d)\s*BR?\s*/?\s*(\d)\s*BA?', r'\1BR/\2BA'),  # 2BR/2BA variants
        (r'^(\d)\s*BED.*(\d)\s*BATH', r'\1BR/\2BA'),  # 2 Bed / 2 Bath
        (r'STUDIO', '0BR/1BA'),
    ]

    for pattern, replacement in patterns:
        s = re.sub(pattern, replacement, s)

    # Ensure format is consistent
    match = re.search(r'(\d)BR/(\d)BA', s)
    if match:
        return f"{match.group(1)}BR/{match.group(2)}BA"

    return s


def parse_rent_roll(df: pd.DataFrame, filename: str) -> Dict[str, Any]:
    """Parse rent roll Excel file."""
    result = {
        "source": {
            "file": filename,
            "extractedAt": datetime.utcnow().isoformat() + "Z",
            "type": "rent_roll",
            "sheet": df.attrs.get("sheet_name", "Source Data"),
            "headerRow": df.attrs.get("header_row_number")
        },
        "units": [],
        "unitMix": [],
        "summary": {},
        "provenance": {
            "sheetName": df.attrs.get("sheet_name", "Source Data"),
            "headerRow": df.attrs.get("header_row_number"),
            "columns": {},
            "summary": {}
        },
        "warnings": [],
        "success": True
    }

    ambiguous_headers = df.attrs.get("ambiguous_headers") or []
    if ambiguous_headers:
        result["warnings"].append(
            "Ambiguous merged header column(s) detected (duplicate label(s): "
            + ", ".join(ambiguous_headers)
            + "); left for candidate review to avoid silently mis-mapping rent columns."
        )

    # Map columns
    column_mapping: Dict[str, Any] = {}
    used_columns = set()
    mappings = {
        'unit_id': ['unit', 'unit #', 'unit no', 'unit number', 'unit id', 'apt', 'apt no', 'apartment', 'apartment no'],
        'unit_type': ['type', 'floorplan', 'floor plan', 'plan', 'bed/bath', 'beds baths', 'br ba', 'br/ba', 'unit type', 'layout'],
        'sqft': ['sf', 'sqft', 'sq ft', 'square feet', 'net rentable sf', 'nrsf', 'size'],
        'market_rent': ['market rent', 'market / asking', 'market asking', 'asking rent', 'quoted rent', 'market'],
        'actual_rent': ['current rent', 'contract rent', 'actual rent', 'in-place rent', 'in place rent', 'tenant rent', 'lease rent', 'monthly rent', 'rent'],
        'status': ['status', 'occupancy', 'occupancy status', 'lease status', 'occupied', 'occ'],
        'tenant': ['tenant', 'resident', 'name'],
        'lease_end': ['lease end', 'expiration', 'expires', 'end date']
    }

    for field, keywords in mappings.items():
        column = find_column(list(df.columns), keywords, used_columns)
        if column is not None:
            column_mapping[field] = column
            used_columns.add(column)
            result["provenance"]["columns"][field] = {
                "header": str(column),
                "column": excel_column_letter(df, column)
            }

    if "unit_id" not in column_mapping:
        result["success"] = False
        result["warnings"].append("Could not identify a unit identifier column.")
        return result

    def status_from_row(row: pd.Series) -> str:
        status_col = column_mapping.get("status")
        tenant_col = column_mapping.get("tenant")
        rent_col = column_mapping.get("actual_rent")
        row_number = int(row.get("__excel_row_number", 0) or 0)
        status_text = normalize_text(row.get(status_col)) if status_col is not None else ""
        tenant_text = normalize_text(row.get(tenant_col)) if tenant_col is not None else ""
        actual_rent = clean_currency(row.get(rent_col)) if rent_col is not None else None

        if status_text:
            if re.search(r"\b(vacant|available|unoccupied|unleased|down|offline|model|notice vacant|no)\b", status_text):
                return "vacant"
            if re.search(r"\b(occupied|occ|leased|current|yes|y|true|tenant)\b", status_text) or status_text == "1":
                return "occupied"
            result["warnings"].append(
                f"Ambiguous occupancy status '{row.get(status_col)}' on row {row_number}; inferred from tenant/rent and left for operator review."
            )

        if tenant_text or (actual_rent is not None and actual_rent > 0):
            return "occupied"
        return "vacant"

    # Columns that signal a row carries actual unit data (vs a free-text note).
    data_columns = [
        column_mapping[field]
        for field in ('unit_type', 'sqft', 'market_rent', 'actual_rent', 'status', 'tenant')
        if field in column_mapping
    ]

    def has_unit_data(row: pd.Series) -> bool:
        if not data_columns:
            return True
        for column in data_columns:
            value = row.get(column)
            if pd.isna(value):
                continue
            if str(value).strip():
                return True
        return False

    # Parse each unit
    units = []
    skipped_total_rows = 0
    skipped_blank_rows = 0
    skipped_note_rows = 0
    for _, row in df.iterrows():
        if empty_row(row):
            skipped_blank_rows += 1
            continue

        unit = {}
        unit_id_value = row.get(column_mapping['unit_id'], '')
        unit_id_label = normalize_text(unit_id_value)
        row_number = int(row.get("__excel_row_number", 0) or 0)

        if total_like(unit_id_value) or unit_id_label in {"unit", "unit no", "apt", "apartment"}:
            skipped_total_rows += 1
            continue

        unit['unitId'] = str(unit_id_value).strip()
        if not unit['unitId'] or unit['unitId'].lower() == 'nan':
            skipped_blank_rows += 1
            continue

        # A row that names something in the unit-id column but carries no rent,
        # size, type, status, or tenant data is a trailing note / disclaimer
        # rather than a unit. Exclude it instead of inflating the unit count.
        if not has_unit_data(row):
            skipped_note_rows += 1
            continue

        if 'unit_type' in column_mapping:
            unit['type'] = normalize_unit_type(row.get(column_mapping['unit_type']))

        if 'sqft' in column_mapping:
            unit['sqft'] = clean_currency(row.get(column_mapping['sqft']))

        if 'market_rent' in column_mapping:
            unit['marketRent'] = clean_currency(row.get(column_mapping['market_rent']))

        if 'actual_rent' in column_mapping:
            unit['actualRent'] = clean_currency(row.get(column_mapping['actual_rent']))

        if 'status' in column_mapping:
            unit['status'] = status_from_row(row)
        else:
            unit['status'] = status_from_row(row)

        unit['_sourceRow'] = row_number

        units.append(unit)

    result['units'] = units

    if skipped_total_rows:
        result['warnings'].append(f"Skipped {skipped_total_rows} total/subtotal/header row(s) before aggregation.")
    if skipped_blank_rows:
        result['warnings'].append(f"Skipped {skipped_blank_rows} blank row(s) before aggregation.")
    if skipped_note_rows:
        result['warnings'].append(f"Skipped {skipped_note_rows} trailing note/non-data row(s) before aggregation.")

    # Calculate aggregates
    if units:
        total_units = len(units)
        occupied = sum(1 for u in units if u.get('status') == 'occupied')

        # Unit mix calculation
        type_groups = {}
        for unit in units:
            t = unit.get('type', 'Unknown')
            if t not in type_groups:
                type_groups[t] = {
                    'type': t,
                    'count': 0,
                    'sqft_sum': 0,
                    'market_sum': 0,
                    'actual_sum': 0,
                    'occupied': 0
                }
            type_groups[t]['count'] += 1
            if unit.get('sqft'):
                type_groups[t]['sqft_sum'] += unit['sqft']
            if unit.get('marketRent'):
                type_groups[t]['market_sum'] += unit['marketRent']
            if unit.get('actualRent'):
                type_groups[t]['actual_sum'] += unit['actualRent']
            if unit.get('status') == 'occupied':
                type_groups[t]['occupied'] += 1

        unit_mix = []
        for t, data in type_groups.items():
            unit_mix.append({
                'type': t,
                'count': data['count'],
                'avgSqFt': round(data['sqft_sum'] / data['count']) if data['count'] > 0 else None,
                'marketRent': round(data['market_sum'] / data['count']) if data['count'] > 0 else None,
                'inPlaceRent': round(data['actual_sum'] / data['count']) if data['count'] > 0 else None,
                'occupiedCount': data['occupied']
            })

        result['unitMix'] = sorted(unit_mix, key=lambda x: x['type'])

        # Summary metrics
        total_sqft = sum(u.get('sqft', 0) or 0 for u in units)
        total_market = sum(u.get('marketRent', 0) or 0 for u in units if u.get('marketRent'))
        total_actual = sum(u.get('actualRent', 0) or 0 for u in units if u.get('actualRent'))

        result['summary'] = {
            'totalUnits': total_units,
            'occupiedUnits': occupied,
            'vacantUnits': total_units - occupied,
            'occupancyRate': round(occupied / total_units, 4) if total_units > 0 else 0,
            'totalSqFt': round(total_sqft),
            'avgUnitSqFt': round(total_sqft / total_units) if total_units > 0 else 0,
            'grossPotentialRentMonthly': round(total_market),
            'grossPotentialRentAnnual': round(total_market * 12),
            'inPlaceRentMonthly': round(total_actual),
            'inPlaceRentAnnual': round(total_actual * 12),
            'lossToLeaseMonthly': round(total_market - total_actual),
            'lossToLeaseAnnual': round((total_market - total_actual) * 12),
            'lossToLeasePercent': round((total_market - total_actual) / total_market, 4) if total_market > 0 else 0
        }
        first_data_row = min(u.get('_sourceRow', 0) for u in units if u.get('_sourceRow'))
        last_data_row = max(u.get('_sourceRow', 0) for u in units if u.get('_sourceRow'))
        unit_id_column = column_mapping.get('unit_id')
        status_column = column_mapping.get('status')
        market_column = column_mapping.get('market_rent')
        actual_column = column_mapping.get('actual_rent')
        result['provenance']['summary'] = {
            'totalUnits': excel_location(df, first_data_row, unit_id_column, f"Counted unit rows {first_data_row}-{last_data_row}"),
            'occupancyRate': excel_location(df, first_data_row, status_column, f"Interpreted occupancy for unit rows {first_data_row}-{last_data_row}"),
            'grossPotentialRentAnnual': excel_location(df, first_data_row, market_column, f"Annualized market rent for unit rows {first_data_row}-{last_data_row}"),
            'inPlaceRentAnnual': excel_location(df, first_data_row, actual_column, f"Annualized in-place rent for unit rows {first_data_row}-{last_data_row}"),
            'lossToLeaseAnnual': excel_location(df, first_data_row, actual_column, f"Market rent less in-place rent for unit rows {first_data_row}-{last_data_row}"),
            'unitMix': excel_location(df, first_data_row, column_mapping.get('unit_type'), f"Aggregated unit rows {first_data_row}-{last_data_row} by unit type")
        }

    return result


def parse_t12(df: pd.DataFrame, filename: str) -> Dict[str, Any]:
    """Parse T12 financial statement."""
    result = {
        "source": {
            "file": filename,
            "extractedAt": datetime.utcnow().isoformat() + "Z",
            "type": "t12",
            "sheet": df.attrs.get("sheet_name", "Source Data"),
            "headerRow": df.attrs.get("header_row_number")
        },
        "revenue": {},
        "expenses": {},
        "summary": {},
        "provenance": {
            "sheetName": df.attrs.get("sheet_name", "Source Data"),
            "headerRow": df.attrs.get("header_row_number"),
            "columns": {},
            "summary": {}
        },
        "warnings": [],
        "success": True
    }

    ambiguous_headers = df.attrs.get("ambiguous_headers") or []
    if ambiguous_headers:
        result["warnings"].append(
            "Ambiguous merged header column(s) detected (duplicate label(s): "
            + ", ".join(ambiguous_headers)
            + "); left for candidate review to avoid silently mis-mapping value columns."
        )

    visible_columns = [column for column in df.columns if not str(column).startswith("__")]
    line_item_col = find_column(visible_columns, ['line item', 'account', 'account name', 'description', 'category'])
    if line_item_col is None and visible_columns:
        line_item_col = visible_columns[0]

    # Find total column, usually rightmost with numbers or labeled Total/T12/Annual.
    total_col = find_column(
        list(reversed(visible_columns)),
        ['t12 total', 'trailing 12', 'trailing twelve', 'annual total', 'annualized', 'annual', 'total', 'ytd total']
    )

    if total_col is None:
        numeric_columns = []
        for col in visible_columns:
            if col == line_item_col:
                continue
            numeric_count = sum(1 for value in df[col] if clean_currency(value) is not None)
            if numeric_count > 0:
                numeric_columns.append((numeric_count, col))
        if numeric_columns:
            total_col = numeric_columns[-1][1]

    if total_col is None or line_item_col is None:
        result['success'] = False
        result['warnings'].append("Could not identify line-item and total/annual columns")
        return result

    result["provenance"]["columns"] = {
        "lineItem": {"header": str(line_item_col), "column": excel_column_letter(df, line_item_col)},
        "total": {"header": str(total_col), "column": excel_column_letter(df, total_col)}
    }

    def matching_rows(patterns: List[str]) -> List[tuple[pd.Series, float, str]]:
        matches = []
        for _, row in df.iterrows():
            if empty_row(row):
                continue
            label = normalize_text(row.get(line_item_col))
            if not label:
                continue
            value = clean_currency(row.get(total_col))
            if value is None:
                continue
            if any(re.search(pattern, label) for pattern in patterns):
                matches.append((row, value, label))
        return matches

    def pick_amount(field_name: str, patterns: List[str]) -> Optional[tuple[float, pd.Series]]:
        matches = matching_rows(patterns)
        if len(matches) > 1:
            labels = ", ".join(match[2] for match in matches[:3])
            result["warnings"].append(f"Multiple possible {field_name} rows found ({labels}); first match left for operator review.")
        if not matches:
            return None
        row, value, _ = matches[0]
        return value, row

    revenue_pick = pick_amount('effective gross income', [
        r"\beffective gross (income|revenue)\b",
        r"\btotal (revenue|income)\b"
    ])
    expenses_pick = pick_amount('total operating expenses', [
        r"\btotal (operating )?expenses?\b",
        r"\boperating expense total\b",
        r"\btotal opex\b"
    ])
    noi_pick = pick_amount('net operating income', [
        r"\bnet operating income\b",
        r"\bnoi\b"
    ])

    revenue = {}
    expenses = {}

    for _, row in df.iterrows():
        label = normalize_text(row.get(line_item_col))
        value = clean_currency(row.get(total_col))
        if not label or value is None:
            continue

        if 'gross' in label and 'rent' in label and revenue_pick is None:
            revenue['grossPotentialRent'] = value
        elif 'vacancy' in label:
            revenue['vacancy'] = value
        elif 'concession' in label:
            revenue['concessions'] = value
        elif 'bad debt' in label or 'write' in label:
            revenue['badDebt'] = value
        elif 'other' in label and 'income' in label:
            revenue['otherIncome'] = value
        elif 'tax' in label and 'payroll' not in label:
            expenses['taxes'] = value
        elif 'insurance' in label:
            expenses['insurance'] = value
        elif 'utilit' in label:
            expenses['utilities'] = value
        elif 'repair' in label or 'maintenance' in label:
            expenses['repairs'] = value
        elif 'management' in label:
            expenses['management'] = value
        elif 'payroll' in label or 'personnel' in label or 'salary' in label:
            expenses['payroll'] = value
        elif 'admin' in label:
            expenses['admin'] = value
        elif 'marketing' in label or 'advertising' in label:
            expenses['marketing'] = value
        elif 'contract' in label:
            expenses['contractServices'] = value

    if revenue_pick:
        revenue['effectiveGrossIncome'] = revenue_pick[0]
        result['provenance']['summary']['effectiveGrossIncome'] = excel_location(
            df,
            int(revenue_pick[1].get("__excel_row_number", 0) or 0),
            total_col,
            "Effective gross income / total revenue row"
        )
    if expenses_pick:
        expenses['totalExpenses'] = expenses_pick[0]
        result['provenance']['summary']['totalExpenses'] = excel_location(
            df,
            int(expenses_pick[1].get("__excel_row_number", 0) or 0),
            total_col,
            "Total operating expenses row"
        )
    if noi_pick:
        result['summary']['noi'] = noi_pick[0]
        result['provenance']['summary']['netOperatingIncome'] = excel_location(
            df,
            int(noi_pick[1].get("__excel_row_number", 0) or 0),
            total_col,
            "Net operating income row"
        )

    result['revenue'] = revenue
    result['expenses'] = expenses

    # Calculate summary if we have enough data
    total_rev = revenue.get('effectiveGrossIncome') or revenue.get('totalRevenue', 0)
    total_exp = expenses.get('totalExpenses', 0)
    noi = result['summary'].get('noi') or (total_rev - total_exp if total_rev and total_exp else 0)
    if noi and 'netOperatingIncome' not in result['provenance']['summary']:
        result['warnings'].append("NOI was derived from revenue less expenses because no explicit NOI row was found.")
        source_row = revenue_pick[1] if revenue_pick else expenses_pick[1] if expenses_pick else None
        result['provenance']['summary']['netOperatingIncome'] = excel_location(
            df,
            int(source_row.get("__excel_row_number", 0) or 0) if source_row is not None else None,
            total_col,
            "Derived from sourced revenue and expense rows"
        )

    result['summary'] = {
        'effectiveGrossIncome': total_rev,
        'totalExpenses': total_exp,
        'netOperatingIncome': noi,
        'expenseRatio': round(total_exp / total_rev, 4) if total_rev else 0,
        'noiMargin': round(noi / total_rev, 4) if total_rev else 0
    }

    return result


def main():
    if len(sys.argv) < 2:
        print(json.dumps({
            "error": "Usage: python parse_excel.py <file_path> [--type rent_roll|t12|auto]",
            "success": False
        }))
        sys.exit(1)

    file_path = sys.argv[1]
    doc_type = 'auto'

    if '--type' in sys.argv:
        idx = sys.argv.index('--type')
        if idx + 1 < len(sys.argv):
            doc_type = sys.argv[idx + 1]

    try:
        # Read every sheet without assuming the first row is the header.
        workbook = pd.read_excel(file_path, engine='openpyxl', sheet_name=None, header=None)
        filename = Path(file_path).name

        # Collect merged-range + image metadata once, then forward-fill merged
        # cells so merged headers/labels map correctly (W10). pandas only keeps
        # the top-left value of a merged region.
        sheet_metadata = read_sheet_metadata(file_path)
        for sheet_name, raw_df in list(workbook.items()):
            meta = sheet_metadata.get(sheet_name, {})
            merged_ranges = meta.get("merged") or []
            if merged_ranges:
                workbook[sheet_name] = fill_merged_cells(raw_df, merged_ranges)

        # Detect or use specified type
        if doc_type == 'auto':
            best_sheet_name, best_raw_df = choose_sheet(workbook, 'auto')
            doc_type = detect_document_type(prepare_table(best_raw_df, best_sheet_name), filename)
        else:
            best_sheet_name, best_raw_df = choose_sheet(workbook, doc_type)

        # Image-only / scanned sheet detection (W11). A target sheet that holds
        # embedded image(s) but essentially no extractable tabular text needs OCR
        # and cannot be parsed into fields. Return a clear, non-empty status
        # instead of crashing or silently producing zeroed output.
        target_meta = sheet_metadata.get(best_sheet_name, {})
        image_count = int(target_meta.get("image_count", 0) or 0)
        text_cell_count = int(target_meta.get("text_cell_count", 0) or 0)
        if image_count > 0 and text_cell_count < 3:
            print(json.dumps({
                "success": True,
                "needsOcr": True,
                "status": "needs-ocr",
                "source": {
                    "file": filename,
                    "type": doc_type if doc_type in ("rent_roll", "t12") else "unknown",
                    "sheet": best_sheet_name,
                },
                "warnings": [
                    f"Sheet '{best_sheet_name}' contains {image_count} embedded image(s) and "
                    f"{text_cell_count} text cell(s); it appears to be a scanned/image-only "
                    "document that requires OCR. No tabular fields could be extracted."
                ],
                "fields": [],
            }, indent=2))
            return

        df = prepare_table(best_raw_df, best_sheet_name)

        # Parse based on type
        if doc_type == 'rent_roll':
            result = parse_rent_roll(df, filename)
        elif doc_type == 't12':
            result = parse_t12(df, filename)
        else:
            result = {
                "success": False,
                "error": f"Unknown document type: {doc_type}",
                "detectedType": doc_type
            }

        print(json.dumps(result, indent=2))

    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e),
            "file": file_path
        }))
        sys.exit(1)


if __name__ == '__main__':
    main()
