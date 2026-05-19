#!/usr/bin/env python3
"""Generate committed XLSX parser fixtures used by parser-service tests."""

from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = ROOT / "fixtures" / "parsers"


def save_workbook(workbook: Workbook, file_name: str) -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(FIXTURE_DIR / file_name)


def rent_roll_alternate_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alt Headers"
    sheet.append(["Apt No.", "Floor Plan", "Net Rentable SF", "Market / Asking", "In Place Rent", "Lease Status"])
    sheet.append(["101", "1 Bed / 1 Bath", 720, 1650, 1600, "Current"])
    sheet.append(["102", "1 Bed / 1 Bath", 720, 1650, 0, "Available"])
    sheet.append(["201", "2x2", 1050, 2250, 2200, "Leased"])
    sheet.append(["202", "Studio", 500, 1400, 1375, "Occupied"])
    save_workbook(workbook, "rent-roll-alternate-headers.xlsx")


def rent_roll_totals_and_blanks() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rent Roll"
    sheet.append(["Parkview Apartments", None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    sheet.append(["Unit Number", "Beds/Baths", "Sq Ft", "Market Rent", "Contract Rent", "Occupancy Status"])
    sheet.append(["101", "1BR/1BA", 700, 1600, 1550, "Occupied"])
    sheet.append(["102", "1BR/1BA", 700, 1600, 0, "Vacant"])
    sheet.append([None, None, None, None, None, None])
    sheet.append(["201", "2BR/2BA", 1050, 2250, 2150, "Occupied"])
    sheet.append(["202", "2BR/2BA", 1050, 2250, 2200, "Occupied"])
    sheet.append(["Total", None, 3500, 7700, 5900, None])
    save_workbook(workbook, "rent-roll-totals-and-blanks.xlsx")


def rent_roll_occupancy_conventions() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Occupancy"
    sheet.append(["Apartment", "Layout", "NRSF", "Quoted Rent", "Monthly Rent", "Resident", "Occ?"])
    sheet.append(["301", "1x1", 710, 1700, 1680, "Jordan Lee", "Yes"])
    sheet.append(["302", "1x1", 710, 1700, 0, None, "No"])
    sheet.append(["303", "2 Bed / 2 Bath", 1080, 2350, 2300, "Morgan Chen", "Notice"])
    sheet.append(["304", "2 Bed / 2 Bath", 1080, 2350, 2325, "Taylor Brooks", None])
    save_workbook(workbook, "rent-roll-occupancy-conventions.xlsx")


def rent_roll_messy_realistic() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RR Export"
    sheet.append(["Lakeside Flats - Rent Roll Export", None, None, None, None, None, None, None, None])
    sheet.append(["As of", "2026-04-30", None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None, None])
    sheet.append(["Unit #", "Floor Plan", "NRSF", "Market Rent", "Lease Rent", "Resident", "Occ", "Lease End", "Notes"])
    sheet.append(["101", "1x1", 710, 1725, 1700, "Jordan Lee", "Yes", "2027-03-31", None])
    sheet.append(["102", "1x1", 710, 1725, 0, None, "Available", None, None])
    sheet.append(["103", "Studio", 520, 1450, 1425, "Morgan Chen", "MTM", "2026-06-30", "Month to month"])
    sheet.append([None, None, None, None, None, None, None, None, None])
    sheet.append(["201", "2 Bed / 2 Bath", 1080, 2350, 2300, "Priya Patel", "Notice", "2026-07-31", "Notice given"])
    sheet.append(["202", "2x2", 1080, 2350, 0, None, "Model", None, "Show unit"])
    sheet.append(["301", "3 BR / 2 BA", 1280, 2850, 2800, "Avery Gomez", "Occupied", "2027-01-31", None])
    sheet.append(["302", "3x2", 1280, 2850, 2750, "TBD Transfer", None, "2026-12-31", "Transfer pending"])
    sheet.append(["Grand Total", None, 6660, 15300, 10975, None, None, None, None])
    save_workbook(workbook, "rent-roll-messy-realistic.xlsx")


def t12_multi_sheet() -> None:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover.append(["Parkview Apartments"])
    cover.append(["Workbook includes notes and trailing twelve financials."])
    cover.append(["Prepared for parser fixture coverage."])

    sheet = workbook.create_sheet("Trailing 12")
    sheet.append(["Parkview Apartments T12", None, None, None, None, None, None, None, None, None, None, None, None, None])
    sheet.append(["Account Name", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Annual"])
    sheet.append(["Gross Potential Rent", 150000, 150000, 150000, 150000, 150000, 150000, 150000, 150000, 150000, 150000, 150000, 150000, 1800000])
    sheet.append(["Vacancy Loss", -20000, -20000, -20000, -20000, -20000, -20000, -20000, -20000, -20000, -20000, -20000, -20000, -240000])
    sheet.append(["Effective Gross Income", 130000, 130000, 130000, 130000, 130000, 130000, 130000, 130000, 130000, 130000, 130000, 130000, 1560000])
    sheet.append(["Total Operating Expenses", 51667, 51667, 51667, 51667, 51667, 51667, 51667, 51667, 51667, 51667, 51667, 51663, 620000])
    sheet.append(["Net Operating Income", 78333, 78333, 78333, 78333, 78333, 78333, 78333, 78333, 78333, 78333, 78333, 78337, 940000])
    save_workbook(workbook, "t12-multi-sheet.xlsx")


def t12_messy_realistic() -> None:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Read Me"
    cover.append(["Lakeside Flats trailing financial export"])
    cover.append(["Rows below include owner chart of accounts and annualized totals."])

    sheet = workbook.create_sheet("T12 - Owner Export")
    sheet.append(["Lakeside Flats", None, None, None, None, None, None, None, None, None, None, None, None, None])
    sheet.append(["Trailing twelve through April 2026", None, None, None, None, None, None, None, None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None, None, None, None, None, None, None])
    sheet.append(["Account", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "Annualized"])
    sheet.append(["Gross Potential Rent", 15300, 15300, 15300, 15300, 15300, 15300, 15300, 15300, 15300, 15300, 15300, 15300, 183600])
    sheet.append(["Vacancy Loss", -1800, -1800, -1800, -1850, -1850, -1850, -1850, -1850, -1850, -1850, -1850, -1850, -22000])
    sheet.append(["Concessions", -400, -400, -400, -400, -400, -425, -425, -425, -425, -425, -425, -450, -5000])
    sheet.append(["Effective Gross Income", 13100, 13100, 13100, 13050, 13050, 13025, 13025, 13025, 13025, 13025, 13025, 13000, 156600])
    sheet.append(["Taxes", 1450, 1450, 1450, 1450, 1450, 1450, 1450, 1450, 1450, 1450, 1450, 1450, 17400])
    sheet.append(["Insurance", 650, 650, 650, 650, 650, 650, 650, 650, 650, 650, 650, 650, 7800])
    sheet.append(["Repairs & Maintenance", 1100, 950, 1200, 900, 1250, 1000, 1150, 1050, 1150, 1000, 1200, 1050, 13000])
    sheet.append(["Payroll / Admin", 1000, 1000, 1000, 1000, 1000, 1000, 1000, 1000, 1000, 1000, 1000, 1000, 12000])
    sheet.append(["Management Fee", 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 500, 6000])
    sheet.append(["Utilities", 375, 400, 425, 400, 375, 400, 425, 400, 375, 400, 425, 400, 4800])
    sheet.append(["Marketing", 25, 0, 50, 0, 25, 0, 25, 0, 25, 0, 25, 25, 200])
    sheet.append(["Total Operating Expenses", 5100, 4950, 5275, 4900, 5250, 5000, 5200, 5050, 5150, 5000, 5250, 5075, 61200])
    sheet.append(["Net Operating Income", 8000, 8150, 7825, 8150, 7800, 8025, 7825, 7975, 7875, 8025, 7775, 7925, 95400])
    save_workbook(workbook, "t12-messy-realistic.xlsx")


def main() -> None:
    rent_roll_alternate_headers()
    rent_roll_totals_and_blanks()
    rent_roll_occupancy_conventions()
    rent_roll_messy_realistic()
    t12_multi_sheet()
    t12_messy_realistic()


if __name__ == "__main__":
    main()
