#!/usr/bin/env python3
"""
Deterministic synthetic CRE benchmark dataset generator (EVAL-PLAN Phase 1).

Produces the 3 benchmark synthetic commercial real estate deals under
``eval/benchmark/deals/<dealId>/`` (one per archetype; see ``all_specs()``). For
each deal it emits:

  * ``deal.json``            - deal input, valid against config/deal-schema.json
  * ``documents/*``          - synthetic source docs (rent roll, T12, offering memo)
                               carrying realistic messiness + planted issues
  * ``ground-truth.json``    - the machine-readable answer key (the TRUE planted
                               economics), shaped per EVAL-PLAN section 0(a)

HONESTY MODEL (non-negotiable)
==============================
There is exactly ONE canonical spec per deal (a ``DealSpec``). Every artifact -
the deal.json, the documents, AND the ground truth - is DERIVED FROM that single
spec. Ground truth is computed straight from the planted economic values; it is
NEVER reverse-engineered to match parser output. The documents are generated
FROM the same spec, so docs and truth are consistent by construction.

The rent-roll / T12 / occupancy figures in ground truth are computed with the
SAME aggregation the deterministic parser uses (occupancy = occupied/total,
GPR_annual = sum(market_rent)*12, in_place_annual = sum(actual_rent)*12) because
that aggregation is itself the objectively-correct arithmetic on the planted unit
rows - not because we are chasing the parser's number.

DETERMINISM
===========
No ``random``, no ``datetime.now()``. All values are hardcoded or computed.
PDFs are written with reportlab ``invariant=1`` so the bytes (and file hashes)
are stable run to run. Re-running the generator reproduces byte-identical files.

Usage::

    python eval/generators/generate_deals.py

See ``eval/generators/README.md`` for the reference underwriting model, the
messiness catalog, and the planted-issue catalog.
"""

from __future__ import annotations

import datetime
import io
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

# Fixed epoch stamped into every generated Excel workbook so the .xlsx bytes -
# and therefore the committed file hashes - are byte-identical run to run. An
# .xlsx is a zip; without freezing BOTH the docProps core created/modified
# timestamps AND every zip entry's mtime, openpyxl bakes the wall-clock time in
# and the file hash changes every run (breaking reproducibility).
FIXED_DT = datetime.datetime(2026, 1, 1, 0, 0, 0)
FIXED_ZIP_DATE = (1980, 1, 1, 0, 0, 0)  # earliest valid zip timestamp


def save_xlsx_deterministic(wb: openpyxl.Workbook, out_path: Path) -> None:
    """Save a workbook to byte-stable .xlsx bytes.

    1. Freeze the document core properties (created/modified/creator).
    2. Repackage the zip rewriting every entry's mtime to a fixed date and
       re-emitting entries in a stable (sorted) order with fixed compression.
    """
    wb.properties.created = FIXED_DT
    wb.properties.modified = FIXED_DT
    wb.properties.creator = "generate_deals.py"
    wb.properties.lastModifiedBy = "generate_deals.py"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fixed_iso = FIXED_DT.strftime("%Y-%m-%dT%H:%M:%SZ")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(buffer) as src:
        names = sorted(src.namelist())
        with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for name in names:
                data = src.read(name)
                if name == "docProps/core.xml":
                    # openpyxl re-stamps <dcterms:modified> with the wall clock at
                    # save time regardless of wb.properties.modified, so normalize
                    # both created and modified to the fixed epoch here.
                    text = data.decode("utf-8")
                    text = re.sub(
                        r"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                        lambda m: m.group(1) + fixed_iso + m.group(2),
                        text,
                    )
                    data = text.encode("utf-8")
                info = zipfile.ZipInfo(filename=name, date_time=FIXED_ZIP_DATE)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                dst.writestr(info, data)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DEALS_ROOT = REPO_ROOT / "eval" / "benchmark" / "deals"


# ===========================================================================
# Reference underwriting model (transparent + documented in README)
# ===========================================================================
#
# Determinable metrics are exact arithmetic on the planted spec. Model-dependent
# metrics (IRR, equity multiple) use the canonical assumptions below, mirroring
# the repo's own workpaper conventions so they are reproducible and defensible.

REV_GROWTH = 0.03          # annual revenue growth
EXP_GROWTH = 0.024         # annual expense growth
EXIT_CAP_SPREAD = 0.0025   # exit cap = going-in cap + 25 bps
EXIT_CAP_FLOOR = 0.055     # realistic market exit-cap floor (see note below)
DEFAULT_HOLD = 5           # years
DEFAULT_CLOSING_COST_PCT = 0.03  # closing costs as % of price when not stated
SALE_COST_PCT = 0.02       # cost of sale at exit (broker + transfer)

# NOTE on EXIT_CAP_FLOOR: EVAL-PLAN defines the reference exit cap as
# "going-in cap + 25 bps". For STABILIZED deals the in-place going-in cap is a
# fair market cap, so that rule applies directly. For value-add / distressed
# deals the IN-PLACE NOI is intentionally depressed, so a literal "going-in cap"
# is artificially low and capitalizing the stabilized exit NOI at it would
# fabricate an absurd (50x) exit value. A real buyer would never exit a
# repositioned asset at a sub-market cap, so the reference model floors the exit
# cap at a defensible market level (5.5%). This keeps the model-dependent
# IRR/EM in a credible band. The floor only binds on the depressed deals and is
# documented in eval/generators/README.md.


def annual_debt_service(loan: float, rate: float, amort_years: int,
                        io_years: int = 0, hold_years: int = DEFAULT_HOLD) -> float:
    """Standard level-payment mortgage annual debt service.

    If the interest-only period covers the entire hold, debt service is pure
    interest (loan * rate). Otherwise the standard amortizing payment formula on
    a monthly basis is annualized. ``amort_years == 0`` is treated as
    interest-only.
    """
    if amort_years <= 0 or io_years >= hold_years:
        return loan * rate
    monthly_rate = rate / 12.0
    n = amort_years * 12
    if monthly_rate == 0:
        return loan / amort_years
    payment = loan * (monthly_rate * (1 + monthly_rate) ** n) / ((1 + monthly_rate) ** n - 1)
    return payment * 12.0


def remaining_balance(loan: float, rate: float, amort_years: int,
                      years_elapsed: int, io_years: int = 0) -> float:
    """Loan balance after ``years_elapsed`` years of amortization.

    Interest-only years do not amortize principal. ``amort_years == 0`` means the
    loan never amortizes (full IO), so the balance stays at the original loan.
    """
    if amort_years <= 0:
        return loan
    monthly_rate = rate / 12.0
    n = amort_years * 12
    amortizing_years = max(0, years_elapsed - io_years)
    months_paid = amortizing_years * 12
    if months_paid <= 0 or monthly_rate == 0:
        return loan
    payment = loan * (monthly_rate * (1 + monthly_rate) ** n) / ((1 + monthly_rate) ** n - 1)
    balance = loan * (1 + monthly_rate) ** months_paid - payment * (
        ((1 + monthly_rate) ** months_paid - 1) / monthly_rate
    )
    return max(0.0, balance)


def irr_bisection(cash_flows: List[float], lo: float = -0.9, hi: float = 1.0,
                  iterations: int = 200) -> float:
    """Solve IRR via bisection on the NPV(rate) sign change.

    ``cash_flows[0]`` is the (negative) initial equity outlay; subsequent entries
    are the period cash flows. Deterministic and dependency-free.
    """
    def npv(rate: float) -> float:
        return sum(cf / (1 + rate) ** i for i, cf in enumerate(cash_flows))

    f_lo = npv(lo)
    f_hi = npv(hi)
    if f_lo * f_hi > 0:
        # No sign change in the bracket; widen once, else return NaN-ish 0.
        hi = 5.0
        f_hi = npv(hi)
        if f_lo * f_hi > 0:
            return float("nan")
    for _ in range(iterations):
        mid = (lo + hi) / 2.0
        f_mid = npv(mid)
        if abs(f_mid) < 1e-6:
            return mid
        if f_lo * f_mid < 0:
            hi = mid
            f_hi = f_mid
        else:
            lo = mid
            f_lo = f_mid
    return (lo + hi) / 2.0


def reference_irr_and_em(year1_noi: float, going_in_cap: float, equity: float,
                         loan: float, rate: float, amort_years: int,
                         io_years: int, hold_years: int = DEFAULT_HOLD,
                         pro_forma_year1_noi: Optional[float] = None) -> Tuple[float, float]:
    """Compute reference levered IRR + equity multiple under canonical assumptions.

    The NOI path is built by ``_grow_noi``: revenue grows at REV_GROWTH and
    expenses at EXP_GROWTH independently (off an implied 55% expense ratio) so the
    documented growth spread is honored rather than collapsed to a single blended
    rate. The exit value uses exit cap = going-in cap + 25 bps (floored) applied
    to the forward (year hold+1) NOI.

    Returns (irr, equity_multiple). ``pro_forma_year1_noi`` is an optional
    starting NOI override for extension/experimentation; the dataset always
    passes ``None`` so the answer key reflects the conservative IN-PLACE NOI
    trajectory (see DealSpec.reference_irr_em).
    """
    start_noi = pro_forma_year1_noi if pro_forma_year1_noi is not None else year1_noi
    noi_path = _grow_noi(start_noi, hold_years + 1)

    # Exit cap = going-in cap + 25 bps, floored at a defensible market level so
    # depressed value-add / distressed in-place caps do not fabricate an absurd
    # exit value (see EXIT_CAP_FLOOR note above).
    exit_cap = max(going_in_cap + EXIT_CAP_SPREAD, EXIT_CAP_FLOOR)
    # Exit value = forward NOI (year hold+1) capitalized at the exit cap.
    gross_sale = noi_path[hold_years] / exit_cap
    net_sale = gross_sale * (1 - SALE_COST_PCT)

    ds = annual_debt_service(loan, rate, amort_years, io_years, hold_years)
    payoff = remaining_balance(loan, rate, amort_years, hold_years, io_years)

    cash_flows = [-equity]
    for year in range(1, hold_years + 1):
        cf = noi_path[year - 1] - ds
        if year == hold_years:
            cf += (net_sale - payoff)
        cash_flows.append(cf)

    irr = irr_bisection(cash_flows)
    # When every period (including the levered reversion) is negative there is no
    # rate that drives NPV to zero, so bisection cannot bracket a root and returns
    # NaN. That is the economically-honest "capital is destroyed as structured"
    # outcome for the distressed dealbreaker deals; represent it with a documented
    # deeply-negative floor rather than NaN so the answer key stays machine-clean.
    if irr != irr:  # NaN check
        irr = -0.99
    total_distributions = sum(cf for cf in cash_flows[1:])
    equity_multiple = total_distributions / equity if equity > 0 else 0.0
    return irr, equity_multiple


def _grow_noi(start_noi: float, years: int,
              expense_ratio: float = 0.55) -> List[float]:
    """Return an NOI path of length ``years``.

    NOI = EGI - OpEx. We back out an implied EGI / OpEx from ``start_noi`` using a
    representative 55% expense ratio, grow EGI at REV_GROWTH and OpEx at EXP_GROWTH
    independently, then recompute NOI each year. This honors the documented growth
    spread rather than growing NOI at a single blended rate.
    """
    egi0 = start_noi / (1 - expense_ratio)
    opex0 = egi0 * expense_ratio
    path = []
    for i in range(years):
        egi = egi0 * (1 + REV_GROWTH) ** i
        opex = opex0 * (1 + EXP_GROWTH) ** i
        path.append(egi - opex)
    return path


# ===========================================================================
# Canonical spec data structures
# ===========================================================================


@dataclass
class UnitGroup:
    """A homogeneous block of identical units in the planted rent roll.

    ``count`` units of ``unit_type`` at ``sqft`` each, with the given monthly
    ``market_rent`` and ``in_place_rent``; ``occupied`` of them are occupied.
    Vacant units contribute their market rent to GPR but $0 to in-place rent
    (mirroring how an analyst and the parser treat a vacant line).
    """
    unit_type: str          # canonical label, e.g. "1BR/1BA"
    rr_label: str           # how it appears in the rent roll doc, e.g. "1x1"
    count: int
    sqft: int
    market_rent: int
    in_place_rent: int
    occupied: int           # number occupied (<= count)


@dataclass
class T12Line:
    """A single T12 line item: ``label`` -> annual ``amount`` (USD)."""
    label: str
    amount: float
    category: str  # "revenue" | "expense" | "egi" | "opex_total" | "noi"


@dataclass
class DealSpec:
    deal_id: str
    deal_name: str
    archetype: str          # "core-plus" | "value-add" | "distressed"
    strategy: str           # schema enum: core | core-plus | value-add | opportunistic
    narrative: str

    # Property
    address: str
    city: str
    state: str
    zip_code: str
    county: str
    year_built: int
    unit_groups: List[UnitGroup]

    # Pricing / financing
    asking_price: float
    target_ltv: float
    estimated_rate: float
    loan_term: int
    amortization: int
    loan_type: str
    interest_only: bool
    io_period: int

    # T12 (the canonical income statement; documents are rendered from these)
    t12_lines: List[T12Line]

    # Strategy targets (deal.json sponsor underwriting targets)
    target_hold: int
    target_irr: float
    target_em: float
    target_coc: float

    # Seller / timeline
    seller_entity: str
    seller_name: str
    broker: str
    broker_firm: str
    psa_date: str
    dd_start: str
    dd_expiration: str
    closing_date: str

    # Document plan + planted issues
    documents: List[Dict[str, Any]]
    om_overrides: Dict[str, Any]  # offering-memo headline overrides (e.g. conflicting NOI)
    red_flags: List[Dict[str, Any]]
    dealbreakers: List[Dict[str, Any]]
    ic_verdict: Dict[str, str]

    # Optional pro-forma stabilized NOI for value-add IRR modeling
    pro_forma_noi: Optional[float] = None
    pro_forma_occupancy: Optional[float] = None
    renovation_budget: Optional[float] = None
    notes: str = ""

    # ----- derived helpers (computed from the planted unit rows / T12) -----

    @property
    def total_units(self) -> int:
        return sum(g.count for g in self.unit_groups)

    @property
    def occupied_units(self) -> int:
        return sum(g.occupied for g in self.unit_groups)

    @property
    def in_place_occupancy(self) -> float:
        return round(self.occupied_units / self.total_units, 4)

    @property
    def gpr_monthly(self) -> int:
        # Gross potential rent at 100% occupancy and market rents.
        return sum(g.count * g.market_rent for g in self.unit_groups)

    @property
    def gpr_annual(self) -> int:
        return self.gpr_monthly * 12

    @property
    def in_place_rent_monthly(self) -> int:
        # Sum of actual rents across occupied units (vacant -> $0).
        return sum(g.occupied * g.in_place_rent for g in self.unit_groups)

    @property
    def in_place_rent_annual(self) -> int:
        return self.in_place_rent_monthly * 12

    def t12_value(self, category: str) -> Optional[float]:
        for line in self.t12_lines:
            if line.category == category:
                return line.amount
        return None

    @property
    def egi(self) -> float:
        v = self.t12_value("egi")
        if v is None:
            raise ValueError(f"{self.deal_id}: missing EGI line")
        return v

    @property
    def opex_total(self) -> float:
        v = self.t12_value("opex_total")
        if v is None:
            raise ValueError(f"{self.deal_id}: missing OpEx total line")
        return v

    @property
    def noi(self) -> float:
        v = self.t12_value("noi")
        if v is None:
            return self.egi - self.opex_total
        return v

    @property
    def going_in_cap(self) -> float:
        return self.noi / self.asking_price

    @property
    def loan_amount(self) -> float:
        return self.asking_price * self.target_ltv

    @property
    def closing_costs(self) -> float:
        return round(self.asking_price * DEFAULT_CLOSING_COST_PCT)

    @property
    def equity(self) -> float:
        return self.asking_price + self.closing_costs - self.loan_amount

    @property
    def annual_debt_service(self) -> float:
        return annual_debt_service(
            self.loan_amount, self.estimated_rate, self.amortization,
            self.io_period if self.interest_only else 0, self.target_hold,
        )

    @property
    def dscr(self) -> float:
        return self.noi / self.annual_debt_service

    def reference_irr_em(self) -> Tuple[float, float]:
        # The reference IRR/EM run on the IN-PLACE NOI trajectory (grown at the
        # documented rev/exp rates), NOT on the sponsor's pro-forma stabilization
        # jump. This is the conservative, defensible buy-hold-grow-sell return on
        # the economics actually evidenced by the documents. proFormaNOI remains
        # in deal.json as the sponsor's underwriting target but does not inflate
        # the answer key (which would otherwise fabricate optimistic returns).
        return reference_irr_and_em(
            year1_noi=self.noi,
            going_in_cap=self.going_in_cap,
            equity=self.equity,
            loan=self.loan_amount,
            rate=self.estimated_rate,
            amort_years=self.amortization,
            io_years=self.io_period if self.interest_only else 0,
            hold_years=self.target_hold,
            pro_forma_year1_noi=None,
        )

    def unit_mix_for_deal_json(self) -> List[Dict[str, Any]]:
        return [
            {
                "type": g.unit_type,
                "count": g.count,
                "avgSqFt": g.sqft,
                "marketRent": g.market_rent,
                "inPlaceRent": g.in_place_rent,
            }
            for g in self.unit_groups
        ]


# ===========================================================================
# deal.json construction
# ===========================================================================


def build_deal_json(spec: DealSpec) -> Dict[str, Any]:
    avg_sqft = round(
        sum(g.count * g.sqft for g in spec.unit_groups) / spec.total_units
    )
    total_sqft = sum(g.count * g.sqft for g in spec.unit_groups)

    financials: Dict[str, Any] = {
        "askingPrice": spec.asking_price,
        "pricePerUnit": round(spec.asking_price / spec.total_units),
        "currentNOI": round(spec.noi),
        "inPlaceOccupancy": spec.in_place_occupancy,
        "trailingT12Revenue": round(spec.egi),
        "trailingT12Expenses": round(spec.opex_total),
        "estimatedClosingCosts": spec.closing_costs,
    }
    if spec.pro_forma_noi is not None:
        financials["proFormaNOI"] = round(spec.pro_forma_noi)
    if spec.pro_forma_occupancy is not None:
        financials["marketOccupancy"] = spec.pro_forma_occupancy
    if spec.renovation_budget is not None:
        financials["renovationBudget"] = spec.renovation_budget

    deal: Dict[str, Any] = {
        "$schema": "../../../../config/deal-schema.json",
        "dealId": spec.deal_id,
        "dealName": spec.deal_name,
        "property": {
            "address": spec.address,
            "city": spec.city,
            "state": spec.state,
            "zip": spec.zip_code,
            "county": spec.county,
            "propertyType": "multifamily",
            "yearBuilt": spec.year_built,
            "totalUnits": spec.total_units,
            "totalSqFt": total_sqft,
            "avgUnitSqFt": avg_sqft,
            "unitMix": {"types": spec.unit_mix_for_deal_json()},
        },
        "financials": financials,
        "financing": {
            "targetLTV": spec.target_ltv,
            "estimatedRate": spec.estimated_rate,
            "loanTerm": spec.loan_term,
            "amortization": spec.amortization,
            "loanType": spec.loan_type,
            "interestOnly": spec.interest_only,
            "ioPeriod": spec.io_period,
        },
        "investmentStrategy": spec.strategy,
        "targetHoldPeriod": spec.target_hold,
        "targetIRR": spec.target_irr,
        "targetEquityMultiple": spec.target_em,
        "targetCashOnCash": spec.target_coc,
        "seller": {
            "name": spec.seller_name,
            "entity": spec.seller_entity,
            "broker": spec.broker,
            "brokerFirm": spec.broker_firm,
        },
        "timeline": {
            "psaExecutionDate": spec.psa_date,
            "ddStartDate": spec.dd_start,
            "ddExpirationDate": spec.dd_expiration,
            "closingDate": spec.closing_date,
        },
    }
    if spec.notes:
        deal["notes"] = spec.notes
    return deal


# ===========================================================================
# Document rendering (rent roll, T12, offering memo) - all from the spec
# ===========================================================================


def _expand_unit_rows(spec: DealSpec) -> List[Tuple[str, str, int, int, int, str]]:
    """Expand unit groups into individual rent-roll rows.

    Returns tuples of (unit_id, rr_label, sqft, market_rent, current_rent, status)
    where current_rent is the in-place rent for occupied units and 0 for vacant.
    Unit numbering is deterministic per group (floor.index style).
    """
    rows: List[Tuple[str, str, int, int, int, str]] = []
    floor = 1
    for g in spec.unit_groups:
        for i in range(g.count):
            unit_id = f"{floor}{(i % 99) + 1:02d}"
            is_occupied = i < g.occupied
            status = "Occupied" if is_occupied else "Vacant"
            current = g.in_place_rent if is_occupied else 0
            rows.append((unit_id, g.rr_label, g.sqft, g.market_rent, current, status))
        floor += 1
    return rows


def render_rent_roll_xlsx(spec: DealSpec, out_path: Path, quirks: List[str]) -> None:
    """Render a rent-roll workbook honoring the requested messiness quirks.

    Quirks supported (matched to the EVAL-PLAN messiness column):
      merged-header        - "Market Rent" header merged across two rent columns
      currency-symbols     - rents written as text like "$1,650"
      trailing-notes       - free-text disclaimer rows after the data
      subtotal-rows        - per-block Subtotal rows + a Grand Total
      alt-headers          - alternate header synonyms (Apt / Floor Plan / etc.)
      occupancy-quirks     - ambiguous status tokens (MTM, Notice)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec_sheet_name(spec, "rent_roll")

    rows = _expand_unit_rows(spec)
    use_currency = "currency-symbols" in quirks
    use_alt = "alt-headers" in quirks
    use_merged = "merged-header" in quirks
    use_subtotals = "subtotal-rows" in quirks
    use_trailing = "trailing-notes" in quirks
    use_occ_quirks = "occupancy-quirks" in quirks

    def money(v: int) -> Any:
        return f"${v:,}" if use_currency else v

    # Title banner (benign - skipped by header detection).
    ws["A1"] = f"{spec.deal_name} - Rent Roll"
    ws.merge_cells("A1:F1")

    # Header row (row 2). Alternate synonyms vs canonical.
    if use_alt:
        headers = ["Apt", "Floor Plan", "SF", "Market Rent", "Current Rent", "Occupancy"]
    else:
        headers = ["Unit", "Unit Type", "SqFt", "Market Rent", "Current Rent", "Status"]

    if use_merged:
        # Merge "Market Rent" across the two rent columns (D2:E2). After the
        # parser's unmerge + forward-fill this yields two identically named
        # rent columns -> a genuine ambiguity the parser must flag.
        ws["A2"] = headers[0]
        ws["B2"] = headers[1]
        ws["C2"] = headers[2]
        ws["D2"] = "Market Rent"
        ws["F2"] = headers[5]
        ws.merge_cells("D2:E2")
    else:
        for col_idx, label in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=label)

    # Optional ambiguous occupancy tokens: rewrite a couple of statuses.
    # Applied to the first occupied unit (-> "MTM") and the first vacant
    # unit (-> "Notice"), both of which the parser treats as ambiguous.
    occ_rewrites_applied = {"MTM": False, "Notice": False}

    def status_token(status: str) -> str:
        if not use_occ_quirks:
            return status
        if status == "Occupied" and not occ_rewrites_applied["MTM"]:
            occ_rewrites_applied["MTM"] = True
            return "MTM"   # month-to-month - occupied, ambiguous token
        if status == "Vacant" and not occ_rewrites_applied["Notice"]:
            occ_rewrites_applied["Notice"] = True
            return "Notice"  # on notice - treated vacant, ambiguous token
        return status

    # Body rows, optionally grouped with subtotals.
    if use_subtotals:
        # Emit per-unit-group blocks, each followed by a Subtotal, then a Grand
        # Total. Subtotal/Grand Total rows are excluded by the parser and must
        # NOT be counted as units in ground truth either.
        idx = 0
        grand_market = 0
        grand_current = 0
        for g in spec.unit_groups:
            block_market = 0
            block_current = 0
            for i in range(g.count):
                unit_id, rr_label, sqft, market, current, status = rows[idx]
                idx += 1
                block_market += market
                block_current += current
                ws.append([unit_id, rr_label, sqft, money(market), money(current), status_token(status)])
            grand_market += block_market
            grand_current += block_current
            ws.append(["Subtotal", None, None, money(block_market), money(block_current), None])
        ws.append(["Grand Total", None, None, money(grand_market), money(grand_current), None])
    else:
        for (unit_id, rr_label, sqft, market, current, status) in rows:
            ws.append([unit_id, rr_label, sqft, money(market), money(current), status_token(status)])

    if use_trailing:
        ws.append([None, None, None, None, None, None])
        ws.append(["Notes:", None, None, None, None, None])
        ws.append(["Rents reflect current billing period; subject to change.", None, None, None, None, None])
        ws.append(["Prepared by property management; not audited.", None, None, None, None, None])

    save_xlsx_deterministic(wb, out_path)


def render_t12_xlsx(spec: DealSpec, out_path: Path, quirks: List[str]) -> None:
    """Render a T12 workbook from the canonical line items.

    Quirks supported:
      currency-symbols  - amounts written as text like "$1,248,000"
      multi-sheet       - a decoy summary sheet precedes the real T12 sheet
      trailing-notes    - free-text notes rows after the statement
      alt-headers       - alternate label synonyms (Total OpEx / EGI / etc.)
    """
    use_currency = "currency-symbols" in quirks
    use_multi = "multi-sheet" in quirks
    use_trailing = "trailing-notes" in quirks
    use_alt = "alt-headers" in quirks

    def money(v: float) -> Any:
        return f"${round(v):,}" if use_currency else round(v)

    wb = openpyxl.Workbook()

    if use_multi:
        # Decoy "Summary" sheet first; the parser scores sheets and should pick
        # the real T12 sheet. The decoy holds non-T12 rollup text.
        decoy = wb.active
        decoy.title = "Summary"
        decoy.append(["Property Summary", None, None])
        decoy.append(["Address", spec.address, None])
        decoy.append(["Units", spec.total_units, None])
        decoy.append(["Reporting Period", "Trailing 12 Months", None])
        ws = wb.create_sheet(spec_sheet_name(spec, "t12"))
    else:
        ws = wb.active
        ws.title = spec_sheet_name(spec, "t12")

    # Leading title banner row (single populated cell - skipped by header logic).
    ws.append([f"{spec.deal_name} - Trailing 12 Operating Statement", None, None])

    if use_alt:
        ws.append(["Description", "Monthly Avg", "Trailing 12"])
    else:
        ws.append(["Account", "Monthly Avg", "Annual Total"])

    egi_label_default = "Effective Gross Income"
    opex_label_default = "Total Operating Expenses"
    noi_label_default = "Net Operating Income"
    if use_alt:
        egi_label_default = "Effective Gross Income"
        opex_label_default = "Total OpEx"
        noi_label_default = "NOI"

    for line in spec.t12_lines:
        label = line.label
        if line.category == "egi" and line.label == "__default__":
            label = egi_label_default
        elif line.category == "opex_total" and line.label == "__default__":
            label = opex_label_default
        elif line.category == "noi" and line.label == "__default__":
            label = noi_label_default
        monthly = line.amount / 12.0
        ws.append([label, money(monthly), money(line.amount)])

    if use_trailing:
        ws.append([None, None, None])
        ws.append(["Notes:", None, None])
        ws.append(["Owner-prepared statement; figures unaudited.", None, None])

    save_xlsx_deterministic(wb, out_path)


def _new_pdf_canvas(path: Path) -> canvas.Canvas:
    # invariant=1 -> deterministic bytes (frozen timestamps + content-derived /ID)
    # so the SHA-256 file hash stays stable run to run.
    c = canvas.Canvas(str(path), pagesize=letter, invariant=1)
    c.setTitle("CRE Offering Memorandum")
    c.setAuthor("Synthetic Benchmark Generator")
    c.setSubject("Deterministic synthetic eval fixture")
    c.setCreator("generate_deals.py")
    c.setProducer("generate_deals.py")
    return c


def render_offering_memo_pdf(spec: DealSpec, out_path: Path) -> None:
    """Render a two-page offering-memo PDF with a real text layer.

    Headline metrics use the exact phrasing the PDF parser recognizes:
      "Offering Price: $X", "N units", "Year Built: YYYY",
      "NN% occupancy", "Net Operating Income: $X".
    OM-vs-T12 conflicts (when planted) are applied via ``spec.om_overrides`` so
    the offering memo headline NOI/occupancy can intentionally disagree with the
    T12 (the conflict the agent must catch).
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    c = _new_pdf_canvas(out_path)

    om = spec.om_overrides or {}
    om_noi = round(om.get("noi", spec.noi))
    om_occ_pct = om.get("occupancy_pct", round(spec.in_place_occupancy * 100))
    om_price = round(om.get("asking_price", spec.asking_price))

    # --- Page 1: headline metrics ---
    text = c.beginText(1 * inch, 10 * inch)
    text.setFont("Helvetica-Bold", 16)
    text.textLine(f"{spec.deal_name} - Offering Memorandum")
    text.setFont("Helvetica", 12)
    text.textLine("")
    text.textLine("Investment Highlights")
    text.textLine(f"Offering Price: ${om_price:,}")
    text.textLine(f"Total Units: {spec.total_units} units")
    text.textLine(f"Year Built: {spec.year_built}")
    text.textLine(f"In-Place Occupancy: {om_occ_pct}% occupancy")
    text.textLine(f"Net Operating Income: ${om_noi:,}")
    text.textLine("")
    location = f"{spec.city}, {spec.state}"
    text.textLine(f"{spec.deal_name} is a {spec.total_units}-unit multifamily community located")
    text.textLine(f"in {location}. {spec.narrative}")
    c.drawText(text)
    c.showPage()

    # --- Page 2: rent roll summary table + offering notes ---
    text = c.beginText(1 * inch, 10 * inch)
    text.setFont("Helvetica-Bold", 14)
    text.textLine("Rent Roll Summary")
    text.setFont("Helvetica", 11)
    text.textLine("")
    text.textLine("Unit Type    Count    Avg SqFt    Market Rent")
    for g in spec.unit_groups:
        text.textLine(
            f"{g.unit_type:<12} {g.count:<8} {g.sqft:<11} {g.market_rent}"
        )
    text.textLine("")
    text.setFont("Helvetica-Bold", 12)
    text.textLine("Offering Notes")
    text.setFont("Helvetica", 11)
    for note_line in om.get("notes", []):
        # wrap simple long lines at ~85 chars deterministically
        for chunk in _wrap(note_line, 85):
            text.textLine(chunk)
    c.drawText(text)
    c.showPage()
    c.save()


def render_offering_memo_md(spec: DealSpec, out_path: Path) -> None:
    """Render a Markdown offering memo (alternative to PDF).

    Currently all 3 deals use the PDF offering memo; this renderer is provided so
    the dataset can be extended with .md memos without changing the contract.
    """
    om = spec.om_overrides or {}
    om_noi = round(om.get("noi", spec.noi))
    om_occ_pct = om.get("occupancy_pct", round(spec.in_place_occupancy * 100))
    om_price = round(om.get("asking_price", spec.asking_price))
    lines = [
        f"# {spec.deal_name} - Offering Memorandum",
        "",
        "## Investment Highlights",
        "",
        f"- Offering Price: ${om_price:,}",
        f"- Total Units: {spec.total_units} units",
        f"- Year Built: {spec.year_built}",
        f"- In-Place Occupancy: {om_occ_pct}% occupancy",
        f"- Net Operating Income: ${om_noi:,}",
        "",
        f"{spec.deal_name} is a {spec.total_units}-unit multifamily community in "
        f"{spec.city}, {spec.state}. {spec.narrative}",
        "",
        "## Rent Roll Summary",
        "",
        "| Unit Type | Count | Avg SqFt | Market Rent |",
        "|---|---|---|---|",
    ]
    for g in spec.unit_groups:
        lines.append(f"| {g.unit_type} | {g.count} | {g.sqft} | {g.market_rent} |")
    if om.get("notes"):
        lines += ["", "## Offering Notes", ""]
        lines += [f"- {n}" for n in om["notes"]]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _wrap(text: str, width: int) -> List[str]:
    """Deterministic greedy word-wrap (no dependency on textwrap defaults)."""
    words = text.split()
    out: List[str] = []
    cur = ""
    for w in words:
        if not cur:
            cur = w
        elif len(cur) + 1 + len(w) <= width:
            cur += " " + w
        else:
            out.append(cur)
            cur = w
    if cur:
        out.append(cur)
    return out or [""]


def spec_sheet_name(spec: DealSpec, kind: str) -> str:
    """Deterministic short sheet name per deal/kind (<= 31 chars for Excel)."""
    if kind == "rent_roll":
        return "Rent Roll"
    if kind == "t12":
        return "T12"
    return "Sheet1"


# ===========================================================================
# ground-truth.json construction
# ===========================================================================


def rel_tol(pct: float) -> Dict[str, Any]:
    return {"type": "relative", "pct": pct}


def abs_tol(abs_val: float) -> Dict[str, Any]:
    return {"type": "absolute", "abs": abs_val}


EXACT = {"type": "exact"}


def build_ground_truth(spec: DealSpec) -> Dict[str, Any]:
    """Assemble the answer key purely from the canonical spec.

    Every extraction field's value is the TRUE planted value computed from the
    spec, tagged with the parser's actual output dot-path and the document it
    lives in. Metrics are computed exactly (determinable) or via the documented
    reference model (model-dependent). Nothing here reads parser output.
    """
    irr, em = spec.reference_irr_em()

    # ---- extraction.fields: only values genuinely present in the docs ----
    fields: List[Dict[str, Any]] = []
    doc_types = {d["type"] for d in spec.documents}

    if "rent_roll" in doc_types:
        fields.append({
            "path": "property.totalUnits", "value": spec.total_units,
            "source": "rent_roll", "tolerance": EXACT,
        })
        fields.append({
            "path": "financials.inPlaceOccupancy", "value": spec.in_place_occupancy,
            "source": "rent_roll", "tolerance": abs_tol(0.01),
        })
        fields.append({
            "path": "financials.grossPotentialRentAnnual", "value": spec.gpr_annual,
            "source": "rent_roll", "tolerance": rel_tol(0.01),
        })
        fields.append({
            "path": "financials.inPlaceRentAnnual", "value": spec.in_place_rent_annual,
            "source": "rent_roll", "tolerance": rel_tol(0.01),
        })

    if "t12" in doc_types:
        fields.append({
            "path": "financials.trailingT12Revenue", "value": round(spec.egi),
            "source": "t12", "tolerance": rel_tol(0.005),
        })
        fields.append({
            "path": "financials.trailingT12Expenses", "value": round(spec.opex_total),
            "source": "t12", "tolerance": rel_tol(0.005),
        })
        fields.append({
            "path": "financials.currentNOI", "value": round(spec.noi),
            "source": "t12", "tolerance": rel_tol(0.01),
        })

    if "offering_memo" in doc_types:
        om = spec.om_overrides or {}
        # The offering memo's headline values are what the PDF parser will read.
        # When a conflict is planted, the OM value differs from the T12 - both
        # are honest planted values in their own document.
        om_price = round(om.get("asking_price", spec.asking_price))
        om_noi = round(om.get("noi", spec.noi))
        om_occ = round(om.get("occupancy_pct", round(spec.in_place_occupancy * 100)) / 100, 4)
        fields.append({
            "path": "financials.askingPrice", "value": om_price,
            "source": "offering_memo", "tolerance": rel_tol(0.005),
        })
        fields.append({
            "path": "property.totalUnits", "value": spec.total_units,
            "source": "offering_memo", "tolerance": EXACT,
        })
        fields.append({
            "path": "property.yearBuilt", "value": spec.year_built,
            "source": "offering_memo", "tolerance": EXACT,
        })
        fields.append({
            "path": "financials.inPlaceOccupancy", "value": om_occ,
            "source": "offering_memo", "tolerance": abs_tol(0.01),
        })
        fields.append({
            "path": "financials.currentNOI", "value": om_noi,
            "source": "offering_memo", "tolerance": rel_tol(0.01),
        })

    # ---- financials.metrics: determinable (exact) + model-dependent (wide) ----
    metrics = [
        {"key": "noi", "value": round(spec.noi),
         "tolerance": rel_tol(0.02), "class": "determinable"},
        {"key": "egi", "value": round(spec.egi),
         "tolerance": rel_tol(0.02), "class": "determinable"},
        {"key": "capRate", "value": round(spec.going_in_cap, 4),
         "tolerance": abs_tol(0.0015), "class": "determinable"},
        {"key": "dscr", "value": round(spec.dscr, 2),
         "tolerance": abs_tol(0.08), "class": "determinable"},
        {"key": "irr", "value": round(irr, 4),
         "tolerance": abs_tol(0.03), "class": "model-dependent"},
        {"key": "equityMultiple", "value": round(em, 2),
         "tolerance": abs_tol(0.20), "class": "model-dependent"},
    ]

    return {
        "dealId": spec.deal_id,
        "archetype": spec.archetype,
        "narrative": spec.narrative,
        "documents": [
            {"file": d["file"], "type": d["type"], "quirks": d.get("quirks", [])}
            for d in spec.documents
        ],
        "extraction": {"fields": fields},
        "financials": {"metrics": metrics},
        "redFlags": spec.red_flags,
        "dealbreakers": spec.dealbreakers,
        "icVerdict": spec.ic_verdict,
        "referenceModel": {
            "holdYears": spec.target_hold,
            "revenueGrowth": REV_GROWTH,
            "expenseGrowth": EXP_GROWTH,
            "exitCap": round(spec.going_in_cap + EXIT_CAP_SPREAD, 4),
            "equity": round(spec.equity),
            "loanAmount": round(spec.loan_amount),
            "annualDebtService": round(spec.annual_debt_service),
            "note": (
                "irr/equityMultiple computed by the documented reference model in "
                "eval/generators/generate_deals.py (see eval/generators/README.md). "
                "Determinable metrics (noi/egi/capRate/dscr) are exact arithmetic "
                "on the planted spec."
            ),
        },
    }


# ===========================================================================
# Deal definitions (8 canonical specs defined; 3 active via all_specs())
# ===========================================================================
#
# All 8 archetype specs remain defined below for extension, but only 3 are
# active in the benchmark (see all_specs(): cp-stabilized-clean,
# va-overlevered-ltv, ds-occupancy-collapse). Re-add any of the others to
# all_specs() to widen the set.
#
# Each spec is internally consistent: the T12 EGI/OpEx/NOI tie to a credible
# operating statement, the going-in cap = NOI / askingPrice lands in a sane band,
# and the planted issue is real (a number that genuinely breaches a threshold or
# a document that genuinely conflicts / is missing).


def std_t12(egi: float, taxes: float, insurance: float, utilities: float,
            repairs: float, management: float, payroll: float, admin: float,
            marketing: float, contract: float,
            egi_label: str = "__default__", opex_label: str = "__default__",
            noi_label: str = "__default__") -> List[T12Line]:
    """Build a standard T12 line-item list with a balancing OpEx total + NOI.

    The individual expense lines are realistic detail; the parser keys off the
    EGI / OpEx-total / NOI summary rows, but the detail lines make the statement
    look real (and let an analyst recompute). OpEx total = sum of detail lines;
    NOI = EGI - OpEx total (so the statement always ties out)."""
    expense_lines = [
        ("Real Estate Taxes", taxes),
        ("Insurance", insurance),
        ("Utilities", utilities),
        ("Repairs & Maintenance", repairs),
        ("Management Fee", management),
        ("Payroll", payroll),
        ("General & Administrative", admin),
        ("Marketing & Advertising", marketing),
        ("Contract Services", contract),
    ]
    opex_total = sum(amount for _, amount in expense_lines)
    noi = egi - opex_total
    lines = [T12Line(egi_label, egi, "egi")]
    for label, amount in expense_lines:
        lines.append(T12Line(label, amount, "expense"))
    lines.append(T12Line(opex_label, opex_total, "opex_total"))
    lines.append(T12Line(noi_label, noi, "noi"))
    return lines


def _docs(rent_quirks: List[str], t12_quirks: List[str],
          om_kind: str = "offering-memo.pdf") -> List[Dict[str, Any]]:
    return [
        {"file": "documents/rent-roll.xlsx", "type": "rent_roll", "quirks": rent_quirks},
        {"file": "documents/t12.xlsx", "type": "t12", "quirks": t12_quirks},
        {"file": f"documents/{om_kind}", "type": "offering_memo", "quirks": []},
    ]


def deal_1_clean() -> DealSpec:
    """#1 cp-stabilized-clean - clean core-plus, NO planted flags (FP control)."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 48, 720, 1650, 1600, occupied=46),
        UnitGroup("2BR/2BA", "2x2", 72, 1050, 2250, 2175, occupied=67),
    ]
    # 113/120 occupied -> 0.9417 occupancy. EGI tuned to a clean 6.0% cap.
    t12 = std_t12(
        egi=2_360_000, taxes=305_000, insurance=92_000, utilities=181_000,
        repairs=148_000, management=94_000, payroll=236_000, admin=46_000,
        marketing=33_000, contract=75_000,
    )
    # NOI = 2,360,000 - 1,210,000 = 1,150,000 ; price 19,170,000 -> cap ~6.00%
    return DealSpec(
        deal_id="cp-stabilized-clean",
        deal_name="Maplewood Commons",
        archetype="core-plus",
        strategy="core-plus",
        narrative="A stabilized, well-occupied community with in-place cash flow and no material defects; included as a false-positive control (no planted red flags).",
        address="4200 Maplewood Drive", city="Austin", state="TX", zip_code="78745",
        county="Travis", year_built=2008, unit_groups=units,
        asking_price=19_170_000, target_ltv=0.65, estimated_rate=0.058,
        loan_term=10, amortization=30, loan_type="Agency",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.13, target_em=1.7, target_coc=0.07,
        seller_entity="Maplewood Holdings LLC", seller_name="R. Patel",
        broker="Jane Cole", broker_firm="Apex Capital Markets",
        psa_date="2026-02-02", dd_start="2026-02-03", dd_expiration="2026-03-05",
        closing_date="2026-04-06",
        documents=_docs(rent_quirks=["alt-headers"], t12_quirks=[]),
        om_overrides={"notes": [
            "Property is professionally managed and 94% occupied with stable collections.",
            "Recent roof and HVAC capital work completed; no deferred maintenance noted.",
        ]},
        red_flags=[],
        dealbreakers=[],
        ic_verdict={
            "value": "PASS", "directional": "go",
            "rationale": "Stabilized core-plus asset; ~6.0% going-in cap, DSCR well above 1.25x, no material red flags. Clean control deal.",
        },
        notes="Stabilized core-plus control deal with no planted issues.",
    )


def deal_2_insurance() -> DealSpec:
    """#2 cp-insurance-understated - T12 insurance line is understated."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 54, 700, 1600, 1560, occupied=51),
        UnitGroup("2BR/2BA", "2x2", 66, 1020, 2200, 2120, occupied=62),
    ]
    # Insurance booked at $41,000 (~$342/unit) - well below the ~$1,000/unit
    # market for this vintage/region; understated by roughly $79,000.
    t12 = std_t12(
        egi=2_280_000, taxes=298_000, insurance=41_000, utilities=176_000,
        repairs=151_000, management=91_000, payroll=228_000, admin=44_000,
        marketing=31_000, contract=72_000,
    )
    return DealSpec(
        deal_id="cp-insurance-understated",
        deal_name="Riverside Gardens",
        archetype="core-plus",
        strategy="core-plus",
        narrative="Stabilized asset whose T12 insurance line is materially below market for the vintage and region; corrected expenses compress NOI and the cap rate.",
        address="1820 Riverside Parkway", city="Dallas", state="TX", zip_code="75215",
        county="Dallas", year_built=2005, unit_groups=units,
        asking_price=20_000_000, target_ltv=0.65, estimated_rate=0.06,
        loan_term=10, amortization=30, loan_type="Agency",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.13, target_em=1.7, target_coc=0.065,
        seller_entity="Riverside Gardens LP", seller_name="M. Alvarez",
        broker="Tom Reed", broker_firm="Lone Star Realty Advisors",
        psa_date="2026-02-09", dd_start="2026-02-10", dd_expiration="2026-03-12",
        closing_date="2026-04-13",
        documents=_docs(rent_quirks=["currency-symbols"], t12_quirks=["currency-symbols", "multi-sheet"]),
        om_overrides={"notes": [
            "Insurance expense reflects the seller's legacy master policy and may not be assumable.",
            "Buyer should obtain a fresh insurance quote during due diligence.",
        ]},
        red_flags=[{
            "id": "insurance-understated", "category": "UNDERWRITING", "severity": "HIGH",
            "required": True,
            "keywords": ["insurance", "understated", "below market", "reassess", "underinsured", "reunderwrite"],
        }],
        dealbreakers=[],
        ic_verdict={
            "value": "CONDITIONAL", "directional": "go",
            "rationale": "Viable core-plus deal but the T12 insurance line is understated; underwriting must reset insurance to market (~$1,000/unit), which compresses NOI. Proceed conditional on re-underwritten expenses.",
        },
        notes="Core-plus with an understated insurance line in the T12.",
    )


def deal_3_concentration() -> DealSpec:
    """#3 cp-concentration-risk - single-employer tenant concentration."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 60, 690, 1500, 1455, occupied=57),
        UnitGroup("2BR/2BA", "2x2", 60, 980, 1980, 1920, occupied=56),
    ]
    t12 = std_t12(
        egi=1_980_000, taxes=246_000, insurance=78_000, utilities=158_000,
        repairs=132_000, management=79_000, payroll=198_000, admin=39_000,
        marketing=27_000, contract=63_000,
    )
    return DealSpec(
        deal_id="cp-concentration-risk",
        deal_name="Foundry Yard Apartments",
        archetype="core-plus",
        strategy="core-plus",
        narrative="Cash-flowing asset whose resident base is heavily concentrated in employees of a single large local employer; a layoff or relocation would spike vacancy.",
        address="55 Foundry Yard Road", city="Greenville", state="SC", zip_code="29605",
        county="Greenville", year_built=2011, unit_groups=units,
        asking_price=17_000_000, target_ltv=0.60, estimated_rate=0.061,
        loan_term=10, amortization=30, loan_type="Agency",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.14, target_em=1.75, target_coc=0.07,
        seller_entity="Foundry Yard Partners LLC", seller_name="D. Nguyen",
        broker="S. Whitfield", broker_firm="Piedmont Multifamily Group",
        psa_date="2026-02-16", dd_start="2026-02-17", dd_expiration="2026-03-19",
        closing_date="2026-04-20",
        documents=_docs(rent_quirks=["trailing-notes"], t12_quirks=[]),
        om_overrides={"notes": [
            "Approximately 60% of residents are employed by Carolina Logistics, the area's largest employer.",
            "The single-employer tenant concentration creates correlated lease-rollover and vacancy risk if that employer downsizes or relocates.",
        ]},
        red_flags=[{
            "id": "tenant-concentration", "category": "MARKET", "severity": "MEDIUM",
            "required": True,
            "keywords": ["concentration", "single employer", "tenant concentration", "employer", "correlated", "Carolina Logistics"],
        }],
        dealbreakers=[],
        ic_verdict={
            "value": "CONDITIONAL", "directional": "go",
            "rationale": "Solid in-place cash flow but material single-employer tenant concentration. Proceed conditional on diversification analysis and a vacancy stress test.",
        },
        notes="Core-plus with single-employer tenant concentration disclosed in the OM.",
    )


def deal_4_dscr_sub120() -> DealSpec:
    """#4 va-sub120-dscr - going-in DSCR ~1.10-1.18 (below 1.20 lender floor)."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 72, 660, 1420, 1280, occupied=63),
        UnitGroup("2BR/2BA", "2x2", 78, 960, 1850, 1690, occupied=68),
    ]
    # In-place NOI modest vs a fairly aggressive 72% LTV at 6.5% -> DSCR ~1.13x.
    # NOI = 2,020,000 - 1,062,000 = 958,000; DS ~846,000 -> DSCR ~1.13 (sub-1.20).
    t12 = std_t12(
        egi=2_020_000, taxes=232_000, insurance=86_000, utilities=171_000,
        repairs=158_000, management=74_000, payroll=205_000, admin=41_000,
        marketing=29_000, contract=66_000,
    )
    return DealSpec(
        deal_id="va-sub120-dscr",
        deal_name="Brookhaven Flats",
        archetype="value-add",
        strategy="value-add",
        narrative="Value-add deal where going-in NOI supports only a ~1.1x DSCR at the targeted leverage, below the typical 1.20x agency floor; cash flow is thin until renovations lift rents.",
        address="900 Brookhaven Avenue", city="Phoenix", state="AZ", zip_code="85021",
        county="Maricopa", year_built=1999, unit_groups=units,
        asking_price=15_500_000, target_ltv=0.72, estimated_rate=0.065,
        loan_term=10, amortization=30, loan_type="Bank",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.16, target_em=1.85, target_coc=0.06,
        seller_entity="Brookhaven Flats LLC", seller_name="K. Osei",
        broker="L. Marsh", broker_firm="Desert Capital Advisors",
        psa_date="2026-02-23", dd_start="2026-02-24", dd_expiration="2026-03-26",
        closing_date="2026-04-27",
        documents=_docs(rent_quirks=["merged-header"], t12_quirks=[]),
        om_overrides={"notes": [
            "Significant loss-to-lease versus market supports a value-add renovation thesis.",
            "Going-in debt service coverage is tight at the targeted leverage and improves only after the renovation lifts rents.",
        ]},
        pro_forma_noi=1_460_000, pro_forma_occupancy=0.94, renovation_budget=1_400_000,
        red_flags=[{
            "id": "dscr-sub-120", "category": "FINANCING", "severity": "HIGH",
            "required": True,
            "keywords": ["dscr", "1.20", "below 1.2", "debt service coverage", "thin coverage", "below lender floor", "1.1"],
        }],
        dealbreakers=[],
        ic_verdict={
            "value": "CONDITIONAL", "directional": "go",
            "rationale": "Renovation upside is real but going-in DSCR is ~1.1x, below the 1.20x lender floor. Proceed conditional on lower leverage, an interest-only period, or a rate buy-down.",
        },
        notes="Value-add with sub-1.20x going-in DSCR at targeted leverage.",
    )


def deal_5_overlevered() -> DealSpec:
    """#5 va-overlevered-ltv - targetLTV 0.82 (over-levered)."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 66, 680, 1480, 1360, occupied=60),
        UnitGroup("2BR/2BA", "2x2", 90, 990, 1900, 1780, occupied=82),
    ]
    t12 = std_t12(
        egi=2_120_000, taxes=268_000, insurance=92_000, utilities=183_000,
        repairs=164_000, management=85_000, payroll=224_000, admin=43_000,
        marketing=31_000, contract=70_000,
    )
    return DealSpec(
        deal_id="va-overlevered-ltv",
        deal_name="Crestline Park",
        archetype="value-add",
        strategy="value-add",
        narrative="Value-add deal underwritten at an aggressive 82% LTV; leverage exceeds prudent limits and the going-in DSCR is correspondingly strained.",
        address="3100 Crestline Park Way", city="Charlotte", state="NC", zip_code="28208",
        county="Mecklenburg", year_built=2001, unit_groups=units,
        asking_price=18_400_000, target_ltv=0.82, estimated_rate=0.064,
        loan_term=10, amortization=30, loan_type="Bridge",
        interest_only=True, io_period=2,
        t12_lines=t12,
        target_hold=5, target_irr=0.17, target_em=1.9, target_coc=0.06,
        seller_entity="Crestline Park Investors LLC", seller_name="P. Romano",
        broker="H. Bell", broker_firm="Queen City Realty Capital",
        psa_date="2026-03-02", dd_start="2026-03-03", dd_expiration="2026-04-02",
        closing_date="2026-05-04",
        documents=_docs(rent_quirks=["subtotal-rows"], t12_quirks=[]),
        om_overrides={"notes": [
            "Sponsor proposes 82% leverage to maximize equity returns on the value-add plan.",
            "Loan-to-value above 80% materially elevates refinance and interest-rate risk if the business plan slips.",
        ]},
        pro_forma_noi=1_560_000, pro_forma_occupancy=0.93, renovation_budget=1_650_000,
        red_flags=[{
            "id": "over-levered-ltv", "category": "FINANCING", "severity": "HIGH",
            "required": True,
            "keywords": ["ltv", "82%", "over-levered", "overleveraged", "leverage", "above 80", "refinance risk", "0.82"],
        }],
        dealbreakers=[],
        ic_verdict={
            "value": "CONDITIONAL", "directional": "go",
            "rationale": "Attractive value-add basis but 82% LTV is over-levered with elevated refinance risk. Proceed conditional on reducing leverage to <=75% or securing additional equity.",
        },
        notes="Value-add underwritten at an over-levered 82% LTV.",
    )


def deal_6_missing_phase1() -> DealSpec:
    """#6 va-missing-phase1 - no Phase I ESA + conflicting OM-vs-T12 NOI."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 58, 700, 1520, 1410, occupied=52),
        UnitGroup("2BR/2BA", "2x2", 70, 1000, 1950, 1830, occupied=63),
    ]
    # True T12 NOI computed below; the OM headline NOI is intentionally higher
    # (seller-favorable) -> a planted OM-vs-T12 conflict.
    t12 = std_t12(
        egi=2_040_000, taxes=258_000, insurance=84_000, utilities=176_000,
        repairs=149_000, management=82_000, payroll=210_000, admin=41_000,
        marketing=30_000, contract=67_000,
    )
    # NOI(T12) = 2,040,000 - 1,097,000 = 943,000. OM claims 1,030,000.
    return DealSpec(
        deal_id="va-missing-phase1",
        deal_name="Cedar Industrial Lofts",
        archetype="value-add",
        strategy="value-add",
        narrative="Adaptive-reuse multifamily on a former light-industrial parcel with NO Phase I Environmental Site Assessment in the data room (environmental data gap); the offering memo NOI also overstates the T12.",
        address="240 Cedar Industrial Blvd", city="Cleveland", state="OH", zip_code="44113",
        county="Cuyahoga", year_built=1996, unit_groups=units,
        asking_price=15_800_000, target_ltv=0.70, estimated_rate=0.063,
        loan_term=10, amortization=30, loan_type="Bank",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.16, target_em=1.85, target_coc=0.06,
        seller_entity="Cedar Industrial Lofts LLC", seller_name="J. Brandt",
        broker="A. Kowalski", broker_firm="Lakefront Commercial Group",
        psa_date="2026-03-09", dd_start="2026-03-10", dd_expiration="2026-04-09",
        closing_date="2026-05-11",
        documents=_docs(rent_quirks=[], t12_quirks=["trailing-notes"]),
        om_overrides={
            "noi": 1_030_000,  # > true T12 NOI of 943,000 -> conflict
            "notes": [
                "Offering memorandum cites a stabilized NOI that exceeds the trailing-12 operating statement; the discrepancy must be reconciled.",
                "No Phase I Environmental Site Assessment is included in the data room despite the former light-industrial use of the parcel.",
            ],
        },
        pro_forma_noi=1_180_000, pro_forma_occupancy=0.93, renovation_budget=1_300_000,
        red_flags=[
            {
                "id": "missing-phase-1-esa", "category": "ENVIRONMENTAL", "severity": "HIGH",
                "required": True,
                "keywords": ["phase i", "phase 1", "esa", "environmental site assessment", "environmental", "data gap", "missing", "former industrial"],
            },
            {
                "id": "om-t12-noi-conflict", "category": "UNDERWRITING", "severity": "MEDIUM",
                "required": True,
                "keywords": ["noi", "conflict", "discrepancy", "overstated", "om", "offering memo", "reconcile", "differs", "t12"],
            },
        ],
        dealbreakers=[],
        ic_verdict={
            "value": "CONDITIONAL", "directional": "go",
            "rationale": "Workable value-add but two diligence gaps: a missing Phase I ESA on a former industrial parcel and an offering-memo NOI that overstates the T12. Proceed conditional on a Phase I and reconciled NOI.",
        },
        notes="Value-add adaptive reuse missing a Phase I ESA, with an OM NOI that overstates the T12.",
    )


def deal_7_occupancy_collapse() -> DealSpec:
    """#7 ds-occupancy-collapse - in-place occupancy 0.62, no bridge -> DEALBREAKER."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 80, 640, 1300, 1180, occupied=50),
        UnitGroup("2BR/2BA", "2x2", 70, 940, 1700, 1560, occupied=43),
    ]
    # 93/150 occupied = 0.62 occupancy. Distressed in-place cash flow.
    t12 = std_t12(
        egi=1_120_000, taxes=214_000, insurance=83_000, utilities=178_000,
        repairs=176_000, management=56_000, payroll=205_000, admin=44_000,
        marketing=41_000, contract=69_000,
    )
    # NOI = 1,120,000 - 1,066,000 = 54,000 -> ~0.5% cap, deeply distressed.
    return DealSpec(
        deal_id="ds-occupancy-collapse",
        deal_name="Harborview Terrace",
        archetype="distressed",
        strategy="opportunistic",
        narrative="Distressed asset at 62% in-place occupancy with collapsed cash flow; the deal is structured with permanent (non-bridge) financing and no lease-up capital, so debt service cannot be covered - a dealbreaker.",
        address="77 Harborview Terrace", city="Toledo", state="OH", zip_code="43604",
        county="Lucas", year_built=1989, unit_groups=units,
        asking_price=9_800_000, target_ltv=0.70, estimated_rate=0.068,
        loan_term=10, amortization=30, loan_type="Bank",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.20, target_em=2.0, target_coc=0.04,
        seller_entity="Harborview Terrace LLC", seller_name="(REO - lender owned)",
        broker="C. Dunn", broker_firm="Midwest Distressed Assets",
        psa_date="2026-03-16", dd_start="2026-03-17", dd_expiration="2026-04-16",
        closing_date="2026-05-18",
        documents=_docs(rent_quirks=["occupancy-quirks"], t12_quirks=[]),
        om_overrides={"notes": [
            "Property is 62% occupied following deferred maintenance and management turnover.",
            "No bridge or lease-up financing is contemplated; in-place NOI does not cover debt service at the proposed permanent loan terms.",
        ]},
        pro_forma_noi=1_050_000, pro_forma_occupancy=0.90, renovation_budget=2_100_000,
        red_flags=[{
            "id": "deferred-maintenance", "category": "PHYSICAL", "severity": "MEDIUM",
            "required": False,
            "keywords": ["deferred maintenance", "capital", "condition", "repairs"],
        }],
        dealbreakers=[{
            "id": "occupancy-collapse-no-bridge", "required": True,
            "keywords": ["occupancy", "62%", "0.62", "vacant", "lease-up", "no bridge", "cannot cover", "negative cash flow", "collapse"],
        }],
        ic_verdict={
            "value": "FAIL", "directional": "no-go",
            "rationale": "62% occupancy with permanent (non-bridge) financing and no lease-up capital cannot cover debt service. Without a bridge-loan and lease-up restructure this is a no-go.",
        },
        notes="Distressed: 62% occupancy financed with permanent debt and no bridge/lease-up capital.",
    )


def deal_8_dscr_below080() -> DealSpec:
    """#8 ds-dscr-below-080 - going-in DSCR < 0.80 -> DEALBREAKER."""
    units = [
        UnitGroup("1BR/1BA", "1x1", 64, 650, 1350, 1240, occupied=44),
        UnitGroup("2BR/2BA", "2x2", 56, 920, 1720, 1600, occupied=41),
    ]
    # In-place NOI low vs 75% LTV at 7.2% -> going-in DSCR well below 0.80.
    t12 = std_t12(
        egi=1_240_000, taxes=176_000, insurance=74_000, utilities=152_000,
        repairs=149_000, management=62_000, payroll=178_000, admin=38_000,
        marketing=33_000, contract=58_000,
    )
    # NOI = 1,240,000 - 920,000 = 320,000. Loan 0.75*9,200,000=6,900,000 @7.2%
    # amortizing -> DS ~ 561,600; DSCR ~ 0.57 (< 0.80).
    return DealSpec(
        deal_id="ds-dscr-below-080",
        deal_name="Sterling Mill Lofts",
        archetype="distressed",
        strategy="opportunistic",
        narrative="Distressed asset whose in-place NOI produces a going-in DSCR well below 0.80x at the proposed leverage and rate; in-place cash flow cannot service the debt - a dealbreaker.",
        address="412 Sterling Mill Road", city="Birmingham", state="AL", zip_code="35203",
        county="Jefferson", year_built=1992, unit_groups=units,
        asking_price=9_200_000, target_ltv=0.75, estimated_rate=0.072,
        loan_term=10, amortization=30, loan_type="Bank",
        interest_only=False, io_period=0,
        t12_lines=t12,
        target_hold=5, target_irr=0.20, target_em=2.0, target_coc=0.04,
        seller_entity="Sterling Mill Lofts LP", seller_name="(special servicer)",
        broker="V. Crane", broker_firm="Southeast Special Situations",
        psa_date="2026-03-23", dd_start="2026-03-24", dd_expiration="2026-04-23",
        closing_date="2026-05-25",
        documents=_docs(rent_quirks=["currency-symbols", "trailing-notes"], t12_quirks=["currency-symbols"]),
        om_overrides={"notes": [
            "In-place net operating income does not cover debt service at the proposed loan terms.",
            "Going-in debt service coverage ratio is well below 0.80x; the deal cannot be financed as proposed without a substantial price reduction or equity infusion.",
        ]},
        pro_forma_noi=860_000, pro_forma_occupancy=0.90, renovation_budget=1_500_000,
        red_flags=[],
        dealbreakers=[{
            "id": "dscr-sub-080", "required": True,
            "keywords": ["dscr", "0.80", "below 0.8", "debt service coverage", "cannot service", "negative leverage", "0.57"],
        }],
        ic_verdict={
            "value": "FAIL", "directional": "no-go",
            "rationale": "Going-in DSCR is well below 0.80x; in-place NOI cannot service the proposed debt. Not financeable as structured - a no-go absent a major price cut or recapitalization.",
        },
        notes="Distressed: going-in DSCR below 0.80x at proposed leverage/rate.",
    )


def all_specs() -> List[DealSpec]:
    # Full 8-deal benchmark: the 3 determinable-risk deals (1/5/7) PLUS the 5
    # narrative-risk / DSCR-boundary deals (2/3/4/6/8) whose planted issues are
    # buried in the documents (insurance understatement, tenant concentration,
    # sub-1.20 / sub-0.80 DSCR, missing Phase I + OM-vs-T12 NOI conflict). The
    # offline layers run on all 8; the costly live layer is scored on a
    # representative subset that still includes the narrative-risk deals.
    return [
        deal_1_clean(),               # cp-stabilized-clean   (core-plus, PASS, control)
        deal_2_insurance(),           # cp-insurance-understated (CONDITIONAL, narrative)
        deal_3_concentration(),       # cp-concentration-risk (CONDITIONAL, narrative)
        deal_4_dscr_sub120(),         # va-sub120-dscr        (CONDITIONAL, DSCR boundary)
        deal_5_overlevered(),         # va-overlevered-ltv    (CONDITIONAL)
        deal_6_missing_phase1(),      # va-missing-phase1     (CONDITIONAL, narrative)
        deal_7_occupancy_collapse(),  # ds-occupancy-collapse (FAIL, dealbreaker)
        deal_8_dscr_below080(),       # ds-dscr-below-080     (FAIL, dealbreaker)
    ]


# ===========================================================================
# Orchestration
# ===========================================================================


def write_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def doc_quirks(spec: DealSpec, doc_type: str) -> List[str]:
    for d in spec.documents:
        if d["type"] == doc_type:
            return d.get("quirks", [])
    return []


def generate_deal(spec: DealSpec) -> None:
    deal_dir = DEALS_ROOT / spec.deal_id
    docs_dir = deal_dir / "documents"
    docs_dir.mkdir(parents=True, exist_ok=True)

    # deal.json
    write_json(deal_dir / "deal.json", build_deal_json(spec))

    # documents
    render_rent_roll_xlsx(spec, docs_dir / "rent-roll.xlsx", doc_quirks(spec, "rent_roll"))
    render_t12_xlsx(spec, docs_dir / "t12.xlsx", doc_quirks(spec, "t12"))
    # Offering memo: PDF for all 3 (markdown renderer available for extension).
    render_offering_memo_pdf(spec, docs_dir / "offering-memo.pdf")

    # ground-truth.json
    write_json(deal_dir / "ground-truth.json", build_ground_truth(spec))


def main() -> None:
    DEALS_ROOT.mkdir(parents=True, exist_ok=True)
    specs = all_specs()
    seen_ids = set()
    for spec in specs:
        if spec.deal_id in seen_ids:
            raise ValueError(f"Duplicate dealId: {spec.deal_id}")
        seen_ids.add(spec.deal_id)
        generate_deal(spec)
    print(f"Generated {len(specs)} synthetic deals under {DEALS_ROOT}")
    for spec in specs:
        irr, em = spec.reference_irr_em()
        print(
            f"  {spec.deal_id:<26} cap={spec.going_in_cap:.4f} "
            f"dscr={spec.dscr:.2f} irr={irr:.3f} em={em:.2f} "
            f"occ={spec.in_place_occupancy:.4f} verdict={spec.ic_verdict['value']}"
        )


if __name__ == "__main__":
    main()
